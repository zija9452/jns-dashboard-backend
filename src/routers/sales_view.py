from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Optional, List, Dict, Any
import json
from calendar import monthrange

from src.database import get_db
from src.models.user import User
from src.models.invoice import Invoice
from src.models.customer_invoice import CustomerInvoice
from src.models.daily_cash import DailyCash
from src.models.expense import Expense
from src.models.stock_entry import StockEntry, StockEntryType
from src.models.product import Product
from src.models.refund import Refund
from src.models.warehouse_invoice import WarehouseInvoice
from src.auth.session_auth import admin_cashier_employee_required_from_session, admin_employee_required_from_session

router = APIRouter()


@router.get("/dashboard/stats")
async def get_dashboard_stats(
    from_date: str = Query(None, description="Start date in YYYY-MM-DD format (overrides month/year)"),
    to_date: str = Query(None, description="End date in YYYY-MM-DD format (overrides month/year)"),
    month: int = None,
    year: int = None,
    current_user: User = Depends(admin_cashier_employee_required_from_session()),  # All authenticated users can access dashboard
    db: AsyncSession = Depends(get_db)
):
    """
    Get dashboard statistics for a specific date range or month.
    Returns sales, expenses, purchases, stock data, and daily chart data.
    Admin, Cashier, and Employee can all access the dashboard.

    Query params:
    - from_date: Start date in YYYY-MM-DD format (optional, overrides month/year if provided)
    - to_date: End date in YYYY-MM-DD format (optional, overrides month/year if provided)
    - month: Month number (1-12), defaults to current month (used if from_date/to_date not provided)
    - year: Year number (e.g., 2026), defaults to current year (used if from_date/to_date not provided)

    Note: Chart data shows only up to today's date (not future dates)
    Role-based response:
    - Admin/Employee: Full data including chartData
    - Cashier: Only KPI data (no chartData), current date only
    """
    # Get current date
    today = datetime.now().date()

    # Get user role for role-based response
    user_role = current_user.role.name if current_user and current_user.role else "admin"

    # CASHIER: Force current date only, ignore all date parameters
    if user_role == "cashier":
        first_day = today
        data_end_date = today
    elif from_date and to_date:
        # ADMIN/EMPLOYEE: Use date range if provided
        try:
            first_day = datetime.fromisoformat(from_date).date()
            data_end_date = datetime.fromisoformat(to_date).date()
            # Validate date range
            if first_day > data_end_date:
                raise HTTPException(status_code=400, detail="from_date must be before or equal to to_date")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        # ADMIN/EMPLOYEE: Fall back to month/year selection
        if month is None:
            month = today.month
        if year is None:
            year = today.year

        # Validate month and year
        if month < 1 or month > 12:
            raise HTTPException(status_code=400, detail="Invalid month. Must be 1-12")

        # Get first and last day of the month
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])

        # IMPORTANT: Always limit to today's date (past, current, or future months)
        data_end_date = min(last_day, today)

    # Try to get from cache first
    from ..utils.cache import cache
    cache_params = {
        "from_date": str(first_day),
        "to_date": str(data_end_date),
        "user_role": user_role
    }
    cached_result = await cache.get_dashboard_stats(cache_params)
    if cached_result:
        return cached_result
    
    # ==================== SALES CALCULATION ====================
    # Calculate total sales from walk-in invoices (SIN- prefix)
    # Use data_end_date to limit to today for current month
    walkin_statement = select(
        func.sum(Invoice.amount_paid)
    ).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= first_day,
            func.date(Invoice.payment_date) <= data_end_date  # Use data_end_date instead of last_day
        )
    )
    walkin_result = await db.execute(walkin_statement)
    walkin_total_sales = float(walkin_result.scalar_one_or_none() or 0)

    # Calculate total sales from customer invoice payments (CIN- prefix)
    # OPTIMIZED: Filter payments in database using PostgreSQL JSON operators instead of Python loop
    from sqlalchemy import text
    customer_payment_query = text("""
        SELECT COALESCE(SUM((payment.value->>'amount')::numeric), 0) as total_collection
        FROM customer_invoices,
        json_array_elements(customer_invoices.payments_history::json) AS payment(value)
        WHERE customer_invoices.invoice_no LIKE 'CIN-%'
        AND (payment.value->>'date')::date BETWEEN :start_date AND :end_date
    """)
    customer_payment_result = await db.execute(
        customer_payment_query,
        {"start_date": first_day, "end_date": data_end_date}
    )
    customer_total_collection = float(customer_payment_result.scalar_one_or_none() or 0)

    total_sales = walkin_total_sales + customer_total_collection
    
    # ==================== EXPENSES CALCULATION ====================
    expenses_result = await db.execute(
        select(func.sum(Expense.amount)).where(
            and_(
                Expense.expense_date >= first_day,
                Expense.expense_date <= data_end_date  # Use data_end_date
            )
        )
    )
    total_expenses = float(expenses_result.scalar_one_or_none() or 0)
    
    # ==================== PURCHASE/COST CALCULATION ====================
    # Calculate actual cost from walk-in invoices using stored cost_price
    # FIXED: Now uses actual cost_price instead of 70% formula
    from sqlalchemy import text
    walkin_cost_query = text("""
        SELECT COALESCE(SUM(
            (item.value->>'quantity')::numeric *
            (item.value->>'cost_price')::numeric
        ), 0) as total_cost
        FROM invoices,
        json_array_elements(invoices.items::json) AS item(value)
        WHERE invoices.invoice_no LIKE 'SIN-%'
        AND invoices.payment_date::date BETWEEN :start_date AND :end_date
    """)
    walkin_cost_result = await db.execute(
        walkin_cost_query,
        {"start_date": first_day, "end_date": data_end_date}
    )
    total_purchase = float(walkin_cost_result.scalar_one_or_none() or 0)
    
    # ==================== STOCK DATA ====================
    # Combine out of stock and short stock into a SINGLE query
    stock_result = await db.execute(
        select(
            func.count(Product.id).filter(Product.stock_level <= 0).label('out_of_stock'),
            func.count(Product.id).filter(
                and_(
                    Product.stock_level > 0,
                    Product.stock_level < 5
                )
            ).label('short_stock')
        )
    )
    stock_row = stock_result.one()
    out_of_stock = int(stock_row.out_of_stock or 0)
    short_stock = int(stock_row.short_stock or 0)
    
    # ==================== OPENING BALANCE ====================
    # Get opening balance from first day of month
    daily_cash_result = await db.execute(
        select(DailyCash).where(DailyCash.date == first_day)
    )
    daily_cash = daily_cash_result.scalar_one_or_none()
    opening_balance = float(daily_cash.total_opening) if daily_cash else 0.0
    
    # ==================== DAILY CHART DATA ====================
    # Get daily sales and expenses for the date range
    # Use func.date() to get actual dates instead of just day numbers
    
    # Daily walk-in sales using SQL GROUP BY date
    daily_walkin_statement = select(
        func.date(Invoice.payment_date).label('date'),
        func.sum(Invoice.amount_paid).label('total')
    ).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= first_day,
            func.date(Invoice.payment_date) <= data_end_date
        )
    ).group_by(func.date(Invoice.payment_date))

    daily_sales_result = await db.execute(daily_walkin_statement)
    daily_sales = {row.date: float(row.total) for row in daily_sales_result.all()}

    # Daily expenses using SQL GROUP BY date
    daily_expense_statement = select(
        func.date(Expense.expense_date).label('date'),
        func.sum(Expense.amount).label('total')
    ).where(
        and_(
            Expense.expense_date >= first_day,
            Expense.expense_date <= data_end_date
        )
    ).group_by(func.date(Expense.expense_date))

    daily_expense_result = await db.execute(daily_expense_statement)
    daily_expenses = {row.date: float(row.total) for row in daily_expense_result.all()}

    # Generate all dates in range
    from datetime import timedelta
    delta = timedelta(days=1)
    current_date = first_day
    dates = []
    sales_data = []
    expenses_data = []
    
    while current_date <= data_end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        dates.append(date_str)
        sales_data.append(round(daily_sales.get(current_date, 0.0), 2))
        expenses_data.append(round(daily_expenses.get(current_date, 0.0), 2))
        current_date += delta

    # Get admin username and role
    admin_username = current_user.username if current_user else "Admin"
    admin_role = current_user.role.name if current_user and current_user.role else "Admin"

    # Build base response
    response = {
        "totalSales": round(total_sales, 2),
        "totalExpense": round(total_expenses, 2),
        "totalPurchase": round(total_purchase, 2),
        "outOfStock": out_of_stock,
        "shortStock": short_stock,
        "adminUser": admin_username,
        "userRole": admin_role,
        "openingBalance": round(opening_balance, 2),
        "dateRange": {
            "from": first_day.strftime('%Y-%m-%d'),
            "to": data_end_date.strftime('%Y-%m-%d')
        }
    }

    # CASHIER: Do not include chart data (charts only for admin/employee)
    if user_role != "cashier":
        response["chartData"] = {
            "dates": dates,
            "sales": sales_data,
            "expenses": expenses_data
        }
        response["month"] = month
        response["year"] = year
        response["monthName"] = first_day.strftime('%B %Y')

    # Cache the result for 5 minutes
    await cache.set_dashboard_stats(cache_params, response, ttl=300)

    return response


@router.get("/walkin-invoices")
async def get_walkin_invoices(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Get walk-in invoices for the given date range
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Query walk-in invoices (SIN- prefix)
    # Use payment_date for accurate date filtering (matches what user selects)
    statement = select(Invoice).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= from_date_obj,
            func.date(Invoice.payment_date) <= to_date_obj
        )
    )

    result = await db.execute(statement)
    invoices = result.scalars().all()

    # Query refunds for THESE SPECIFIC INVOICES (regardless of refund date)
    # This way, if sale was on April 10 and refund on April 11,
    # selecting April 10 will still show the refund info
    from src.models.refund import Refund
    
    # Get invoice IDs from the query result
    invoice_ids = [inv.id for inv in invoices]
    
    refund_map = {}
    if invoice_ids:  # Only query if we have invoices
        refund_statement = select(Refund).where(
            Refund.invoice_id.in_(invoice_ids)
        )
        refund_result = await db.execute(refund_statement)
        refunds = refund_result.scalars().all()

        # Build a map of invoice_id -> list of refund records (keeping dates separate)
        for refund in refunds:
            invoice_id_str = str(refund.invoice_id)
            if invoice_id_str not in refund_map:
                refund_map[invoice_id_str] = {
                    "is_refunded": True,
                    "refund_records": [
                        {
                            "refund_date": refund.created_at.isoformat(),
                            "refund_amount": float(refund.amount),
                            "refund_reason": refund.reason,
                            "refund_items": json.loads(refund.items) if refund.items else []
                        }
                    ]
                }
            else:
                # Add as a separate refund record (don't merge)
                refund_map[invoice_id_str]["refund_records"].append({
                    "refund_date": refund.created_at.isoformat(),
                    "refund_amount": float(refund.amount),
                    "refund_reason": refund.reason,
                    "refund_items": json.loads(refund.items) if refund.items else []
                })

    invoice_list = []
    for inv in invoices:
        try:
            totals = json.loads(inv.totals) if inv.totals else {}
            items = json.loads(inv.items) if inv.items else []

            total_invoice_amount = float(totals.get('total', 0))
            total_invoice_paid = float(totals.get('amount_paid', 0))
            total_invoice_discount = float(totals.get('discount', 0))

            # Check if this invoice has refunds
            invoice_refund_info = refund_map.get(str(inv.id))

            # Track if this is the first item in the invoice
            first_item = True

            # Create separate row for each item in the invoice
            for item in items:
                product_name = item.get('product_name', item.get('pro_name', 'N/A'))
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                item_total = quantity * unit_price - item_discount

                # Use stored cost_price from invoice item (already saved at invoice creation time)
                # Multiply by quantity to get total cost for this item
                item_cost = float(item.get('cost_price', 0)) * quantity

                # Show payment and discount only on first item
                amount_paid = total_invoice_paid if first_item else 0.0
                item_discount_display = item_discount if first_item else 0.0
                total_discount_display = total_invoice_discount if first_item else 0.0

                invoice_data = {
                    "id": str(inv.id),
                    "invoice_no": inv.invoice_no,
                    "product_name": product_name,
                    "total_amount": float(item_total),
                    "amount_paid": amount_paid,
                    "balance_due": 0.0 if first_item else 0.0,
                    "payment_status": inv.payment_status,
                    "payment_method": inv.payment_method or "cash",
                    "quantity": quantity,
                    "discount": item_discount_display,
                    "total_discount": total_discount_display,
                    "cost": item_cost,
                    "created_at": inv.created_at.isoformat() if inv.created_at else None,
                    # ALL rows get is_refunded=True for red background
                    "is_refunded": invoice_refund_info["is_refunded"] if invoice_refund_info else False,
                    # Only FIRST row gets detailed refund info for tooltip
                    "refund_records": invoice_refund_info["refund_records"] if (invoice_refund_info and first_item) else []
                }

                invoice_list.append(invoice_data)

                first_item = False
        except Exception as e:
            continue

    return {"invoices": invoice_list, "total": len(invoice_list)}


@router.get("/customized-invoices")
async def get_customized_invoices(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Get customized/customer invoices with payments received in the date range.
    
    CASH BASIS approach:
    - Shows invoices that received payments within the selected date range
    - Shows how much was paid on each selected date
    - Simple columns: Total, Today's Payment, Total Paid, Pending
    
    Example:
    - Invoice created on Feb 28
    - Payment 1: Feb 28 (Rs. 10,000) - Shows when Feb 28 selected
    - Payment 2: Mar 3 (Rs. 15,000) - Shows when Mar 3 selected
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Query ALL customer invoices (CIN- prefix)
    # We'll filter by payment date in Python, not SQL
    statement = select(CustomerInvoice).where(
        CustomerInvoice.invoice_no.like("CIN-%")
    )

    result = await db.execute(statement)
    all_invoices = result.scalars().all()

    invoice_list = []
    
    for inv in all_invoices:
        try:
            # Parse payment history
            payment_history = []
            if inv.payments_history:
                try:
                    payment_history = json.loads(inv.payments_history)
                except:
                    payment_history = []
            
            # Filter payments within selected date range
            payments_in_range = []
            for payment in payment_history:
                payment_date_str = payment.get('date', '')
                if payment_date_str:
                    try:
                        # Handle ISO format dates
                        if 'T' in payment_date_str:
                            payment_date = datetime.fromisoformat(payment_date_str).date()
                        else:
                            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                        
                        # Check if payment date is within range
                        if from_date_obj <= payment_date <= to_date_obj:
                            payments_in_range.append({
                                "date": payment_date_str,
                                "amount": float(payment.get('amount', 0)),
                                "method": payment.get('payment_method', 'cash'),
                                "description": payment.get('description', '')
                            })
                    except:
                        continue
            
            # Skip invoices with no payments in the selected date range
            if not payments_in_range:
                continue
            
            # Calculate total payment received in the selected date range
            total_payment_in_range = sum(p['amount'] for p in payments_in_range)
            
            # Get the latest payment time for display
            latest_payment_time = ""
            if payments_in_range:
                latest_payment = payments_in_range[-1]  # Get last payment
                payment_date_str = latest_payment.get('date', '')
                if payment_date_str:
                    try:
                        if 'T' in payment_date_str:
                            payment_datetime = datetime.fromisoformat(payment_date_str)
                            latest_payment_time = payment_datetime.strftime('%I:%M:%S %p')  # Format: 07:16:14 AM
                        else:
                            latest_payment_time = "12:00:00 AM"
                    except:
                        latest_payment_time = "12:00:00 AM"
            
            # Get unique payment methods used in this date range
            methods_in_range = list(set(p['method'] for p in payments_in_range))
            
            # Parse items to get quantity
            items = []
            try:
                items = json.loads(inv.items) if inv.items else []
            except:
                items = []
            
            total_quantity = sum(item.get('quantity', 0) for item in items)
            
            # Calculate values
            total_amount = float(inv.total_amount) if inv.total_amount else 0.0
            total_paid = float(inv.amount_paid) if inv.amount_paid else 0.0
            pending = float(inv.balance_due) if inv.balance_due else 0.0
            
            invoice_list.append({
                "id": str(inv.id),
                "invoice_no": inv.invoice_no,
                "customer_name": inv.customer_name or "N/A",
                "team_name": inv.team_name or "",
                "total_amount": total_amount,
                "payment_in_selected_range": total_payment_in_range,  # Jo payment aaj/selected date mein hui
                "total_paid": total_paid,  # Cumulative total paid till now
                "pending": pending,  # Remaining balance
                "payment_status": inv.payment_status,  # paid/partial/unpaid
                "payment_methods_used": methods_in_range,  # Methods used in selected date
                "quantity": total_quantity,
                "created_at": inv.created_at.isoformat() if inv.created_at else None,
                "payment_time": latest_payment_time,  # Time of payment (e.g., "02:30 PM")
                "payment_at": payments_in_range[-1]['date'] if payments_in_range else (inv.created_at.isoformat() if inv.created_at else None),
                "payments_in_range": payments_in_range  # Detailed payment list for reference
            })
            
        except Exception as e:
            continue

    return {"invoices": invoice_list, "total": len(invoice_list)}


@router.get("/customized-summary")
async def get_customized_sales_summary(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Get summary of customer invoice payments received in the date range.
    
    CASH BASIS: Shows payments received on selected dates, not invoice creation dates.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Query ALL customer invoices
    statement = select(CustomerInvoice).where(
        CustomerInvoice.invoice_no.like("CIN-%")
    )

    result = await db.execute(statement)
    all_invoices = result.scalars().all()

    # Initialize counters
    total_collection = 0.0
    cash_collection = 0.0
    easypaisa_zohaib_collection = 0.0
    easypaisa_yasir_collection = 0.0
    bank_collection = 0.0
    
    invoices_with_payments = 0

    for inv in all_invoices:
        try:
            # Parse payment history
            payment_history = []
            if inv.payments_history:
                try:
                    payment_history = json.loads(inv.payments_history)
                except:
                    payment_history = []
            
            # Filter payments within selected date range
            for payment in payment_history:
                payment_date_str = payment.get('date', '')
                if payment_date_str:
                    try:
                        # Handle ISO format dates
                        if 'T' in payment_date_str:
                            payment_date = datetime.fromisoformat(payment_date_str).date()
                        else:
                            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                        
                        # Check if payment date is within range
                        if from_date_obj <= payment_date <= to_date_obj:
                            amount = float(payment.get('amount', 0))
                            method = payment.get('payment_method', 'cash').lower()
                            
                            total_collection += amount
                            invoices_with_payments += 1
                            
                            # Categorize by payment method
                            if method == 'cash':
                                cash_collection += amount
                            elif 'easypaisa' in method and 'zohaib' in method:
                                easypaisa_zohaib_collection += amount
                            elif 'easypaisa' in method and 'yasir' in method:
                                easypaisa_yasir_collection += amount
                            elif 'bank' in method or 'faysal' in method:
                                bank_collection += amount
                            else:
                                # Default to cash for unknown methods
                                cash_collection += amount
                                
                    except:
                        continue
                        
        except Exception as e:
            continue

    return {
        "total_collection": total_collection,
        "cash": cash_collection,
        "easypaisa_zohaib": easypaisa_zohaib_collection,
        "easypaisa_yasir": easypaisa_yasir_collection,
        "bank": bank_collection,
        "invoices_count": invoices_with_payments
    }


@router.get("/summary")
async def get_sales_summary(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Get combined sales summary for the given date range.
    Includes both walk-in invoices and customer invoice payments.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Get opening balance (from first date in range)
    daily_cash_result = await db.execute(
        select(DailyCash).where(DailyCash.date == from_date_obj)
    )
    daily_cash = daily_cash_result.scalar_one_or_none()
    opening = float(daily_cash.total_opening) if daily_cash else 0.0

    # Calculate WALK-IN SALES directly from Invoice table (NOT DailyCash)
    # Query walk-in invoices (SIN- prefix) for the date range
    walkin_total_sales = 0.0
    walkin_cash_sales = 0.0
    walkin_total_cost = 0.0  # Calculate cost in same loop
    
    walkin_statement = select(Invoice).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= from_date_obj,
            func.date(Invoice.payment_date) <= to_date_obj
        )
    )
    walkin_result = await db.execute(walkin_statement)
    walkin_invoices_for_sales = walkin_result.scalars().all()

    for inv in walkin_invoices_for_sales:
        # Calculate sales
        walkin_total_sales += float(inv.amount_paid)
        if inv.payment_method.lower() == 'cash':
            walkin_cash_sales += float(inv.amount_paid)

        # Calculate cost using stored cost_price from invoice items
        try:
            items = json.loads(inv.items) if inv.items else []
            for item in items:
                quantity = item.get('quantity', 0)
                item_cost = float(item.get('cost_price', 0)) * quantity
                walkin_total_cost += item_cost
        except:
            continue

    # Get customer invoice payments in date range
    customer_statement = select(CustomerInvoice).where(
        CustomerInvoice.invoice_no.like("CIN-%")
    )
    customer_result = await db.execute(customer_statement)
    all_customer_invoices = customer_result.scalars().all()

    customer_total_collection = 0.0
    customer_cash_collection = 0.0

    for inv in all_customer_invoices:
        try:
            payment_history = []
            if inv.payments_history:
                try:
                    payment_history = json.loads(inv.payments_history)
                except:
                    payment_history = []
            
            for payment in payment_history:
                payment_date_str = payment.get('date', '')
                if payment_date_str:
                    try:
                        if 'T' in payment_date_str:
                            payment_date = datetime.fromisoformat(payment_date_str).date()
                        else:
                            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                        
                        if from_date_obj <= payment_date <= to_date_obj:
                            amount = float(payment.get('amount', 0))
                            method = payment.get('payment_method', 'cash').lower()
                            
                            customer_total_collection += amount
                            
                            if method == 'cash':
                                customer_cash_collection += amount
                                
                    except:
                        continue
        except:
            continue

    # Combined totals (walk-in + customer payments)
    total_sales = walkin_total_sales + customer_total_collection

    # Calculate gross profit using ACTUAL walk-in cost ONLY
    # Gross Profit = Total Sales - Total Cost (customer invoices have NO cost tracking)
    gross_profit = total_sales - walkin_total_cost

    # Total purchase/cost = ONLY walk-in cost (customer invoices don't have cost tracking)
    total_purchase = walkin_total_cost

    # Get total cash recovery
    total_cash_sales = walkin_cash_sales + customer_cash_collection

    # Get total expenses
    expenses_result = await db.execute(
        select(func.sum(Expense.amount)).where(
            and_(
                Expense.expense_date >= from_date_obj,
                Expense.expense_date <= to_date_obj
            )
        )
    )
    total_expenses = float(expenses_result.scalar_one_or_none() or 0)

    # Calculate net cash
    total_cash_sales = walkin_cash_sales + customer_total_collection

    net_cash = opening + total_cash_sales - total_expenses

    # Calculate net profit = Gross Profit - Expenses
    net_profit = gross_profit - total_expenses
    
    # Calculate total refund amount for the date range (Walk-in invoices only - SIN- prefix)
    total_refund_amount = 0.0
    
    try:
        # Query refunds for walk-in invoices only (SIN- prefix)
        from ..models.refund import Refund
        # Convert to datetime for proper comparison
        from_date_datetime = datetime.combine(from_date_obj, datetime.min.time())
        to_date_datetime = datetime.combine(to_date_obj, datetime.max.time())
        
        # Join refunds with invoices and filter by SIN- prefix
        refunds_statement = select(Refund).join(
            Invoice, Refund.invoice_id == Invoice.id
        ).where(
            and_(
                Refund.created_at >= from_date_datetime,
                Refund.created_at <= to_date_datetime,
                Invoice.invoice_no.like('SIN-%')  # Only walk-in invoices
            )
        )
        refunds_result = await db.execute(refunds_statement)
        refunds = refunds_result.scalars().all()
        
        for refund in refunds:
            try:
                total_refund_amount += float(refund.amount)
            except:
                continue
    except Exception as e:
        total_refund_amount = 0.0

    return {
        "opening": opening,
        "totalSale": total_sales,
        "grossProfit": gross_profit,
        "totalExpense": total_expenses,
        "totalRecovery": total_cash_sales,
        "vendorPayments": 0.0,
        "netCash": net_cash,
        "totalPurchase": total_purchase,  # Now actual cost from walk-in invoices
        "totalRefund": total_refund_amount,
        "netProfit": net_profit,
        # Breakdown for reference
        "walkin_sales": walkin_total_sales,
        "customer_payments": customer_total_collection,
        "walkin_cost": walkin_total_cost  # Actual cost for reference
    }


# ==================== REPORT ENDPOINTS ====================

@router.get("/walkin-invoices/pdf")
async def get_walkin_invoices_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for walk-in invoices (date-wise)
    """
    import base64
    from datetime import datetime
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query walk-in invoices
    # Use payment_date for accurate date filtering (matches what user selects)
    statement = select(Invoice).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= from_date_obj,
            func.date(Invoice.payment_date) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    invoices = result.scalars().all()

    # Build invoice rows for PDF
    invoice_rows = ""
    total_amount = 0.0
    total_cost = 0.0
    
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            totals = json.loads(inv.totals) if inv.totals else {}

            for item in items:
                product_name = item.get('product_name', item.get('pro_name', 'N/A'))
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                item_total = quantity * unit_price - item_discount
                item_cost = item_total * 0.7

                invoice_rows += f"""
                <tr>
                    <td class="border">{inv.invoice_no}</td>
                    <td class="border">{inv.created_at.strftime('%Y-%m-%d') if inv.created_at else ''}</td>
                    <td class="border">{product_name}</td>
                    <td class="border text-right">{quantity}</td>
                    <td class="border text-right">{unit_price:.0f}</td>
                    <td class="border text-right">{item_discount:.0f}</td>
                    <td class="border text-right">{item_total:.0f}</td>
                    <td class="border text-right">{item_cost:.0f}</td>
                    <td class="border">{inv.created_at.strftime('%I:%M %p') if inv.created_at else ''}</td>
                </tr>
                """
                total_amount += item_total
                total_cost += item_cost
        except:
            continue
    
    # Calculate totals
    gross_profit = total_amount - total_cost
    
    # Create HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 15mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 11px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 24px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 15px 0;
                color: #666;
                font-size: 12px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .border {{
                border: 1px solid #000;
                padding: 6px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 8px;
                font-size: 11px;
            }}
            .text-right {{
                text-align: right;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
            }}
            .summary {{
                margin-top: 20px;
                border: 2px solid #333;
                padding: 10px;
            }}
            .summary-row {{
                display: flex;
                justify-content: space-between;
                margin-bottom: 5px;
            }}
            .summary-label {{
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <h1>Walk-in Invoice Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 12%;">Invoice No</th>
                    <th style="width: 12%;">Date</th>
                    <th style="width: 23%;">Product</th>
                    <th style="width: 7%;">Qty</th>
                    <th style="width: 9%;">Price</th>
                    <th style="width: 9%;">Discount</th>
                    <th style="width: 11%;">Total</th>
                    <th style="width: 11%;">Cost</th>
                    <th style="width: 10%;">Time</th>
                </tr>
            </thead>
            <tbody>
                {invoice_rows}
                <tr class="total-row">
                    <td class="border" colspan="3" style="text-align: left; font-weight: bold;">TOTAL</td>
                    <td class="border"></td>
                    <td class="border"></td>
                    <td class="border"></td>
                    <td class="border text-right" style="font-weight: bold;">{total_amount:.0f}</td>
                    <td class="border text-right" style="font-weight: bold;">{total_cost:.0f}</td>
                    <td class="border"></td>
                </tr>
            </tbody>
        </table>
        
        <div class="summary">
            <div class="summary-row">
                <span class="summary-label">Total Sales:</span>
                <span>Rs. {total_amount:.0f}</span>
            </div>
            <div class="summary-row">
                <span class="summary-label">Total Cost:</span>
                <span>Rs. {total_cost:.0f}</span>
            </div>
            <div class="summary-row">
                <span class="summary-label">Gross Profit:</span>
                <span>Rs. {gross_profit:.0f}</span>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception as e:
        # Fallback - return simple PDF
        pdf_content = f"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\ntrailer\n<< /Size 4 /Root 1 0 R >>\n%%EOF"
        encoded_pdf = base64.b64encode(pdf_content.encode()).decode()

    return {"pdf": encoded_pdf}


@router.get("/walkin-invoices/excel")
async def get_walkin_invoices_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for walk-in invoices (date-wise)
    Returns actual .xlsx file
    """
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query walk-in invoices
    # Use payment_date for accurate date filtering (matches what user selects)
    statement = select(Invoice).where(
        and_(
            Invoice.invoice_no.like("SIN-%"),
            func.date(Invoice.payment_date) >= from_date_obj,
            func.date(Invoice.payment_date) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    invoices = result.scalars().all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Walk-in Invoices"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    headers = [
        'Invoice No',
        'Date',
        'Product Name',
        'Quantity',
        'Unit Price',
        'Discount',
        'Total Amount',
        'Cost',
        'Time'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [15, 12, 30, 10, 12, 12, 15, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    total_amount = 0.0
    total_cost = 0.0
    row_num = 2
    
    # Write data rows
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            
            for item in items:
                product_name = item.get('product_name', item.get('pro_name', 'N/A'))
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                item_total = quantity * unit_price - item_discount
                item_cost = item_total * 0.7
                
                # Write row data
                ws.cell(row=row_num, column=1, value=inv.invoice_no).border = thin_border
                ws.cell(row=row_num, column=2, value=inv.created_at.strftime('%Y-%m-%d') if inv.created_at else '').border = thin_border
                ws.cell(row=row_num, column=3, value=product_name).border = thin_border
                ws.cell(row=row_num, column=4, value=quantity).border = thin_border
                ws.cell(row=row_num, column=5, value=round(unit_price, 2)).border = thin_border
                ws.cell(row=row_num, column=6, value=round(item_discount, 2)).border = thin_border
                ws.cell(row=row_num, column=7, value=round(item_total, 2)).border = thin_border
                ws.cell(row=row_num, column=8, value=round(item_cost, 2)).border = thin_border
                ws.cell(row=row_num, column=9, value=inv.created_at.strftime('%I:%M %p') if inv.created_at else '').border = thin_border
                
                # Apply alignments
                for col in range(1, 10):
                    if col in [4]:  # Quantity
                        ws.cell(row=row_num, column=col).alignment = cell_alignment
                    elif col in [5, 6, 7, 8]:  # Numeric columns
                        ws.cell(row=row_num, column=col).alignment = right_alignment
                    else:
                        ws.cell(row=row_num, column=col).alignment = cell_alignment
                
                total_amount += item_total
                total_cost += item_cost
                row_num += 1
        except:
            continue
    
    # Write total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(total_amount, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=round(total_cost, 2)).font = Font(bold=True)
    
    # Merge cells for TOTAL label
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    
    # Apply border to total row
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).border = thin_border
    
    # Add summary section
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value='Total Sales:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=round(total_amount, 2)).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1, value='Total Cost:').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=round(total_cost, 2)).font = Font(bold=True)
    
    ws.cell(row=summary_row + 2, column=1, value='Gross Profit:').font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=round(total_amount - total_cost, 2)).font = Font(bold=True)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()
    
    # Encode to base64
    encoded_excel = base64.b64encode(excel_bytes).decode()
    
    return {
        "excel": encoded_excel,
        "filename": f"walkin_invoices_{from_date}_to_{to_date}.xlsx"
    }


@router.get("/customized-invoices/pdf")
async def get_customized_invoices_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for customer invoices (date-wise payments)
    """
    import base64
    from datetime import datetime
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query customer invoices
    statement = select(CustomerInvoice).where(
        CustomerInvoice.invoice_no.like("CIN-%")
    )
    result = await db.execute(statement)
    all_invoices = result.scalars().all()
    
    # Build invoice rows for PDF
    invoice_rows = ""
    total_collection = 0.0
    invoice_count = 0
    
    for inv in all_invoices:
        try:
            # Parse payment history
            payment_history = []
            if inv.payments_history:
                try:
                    payment_history = json.loads(inv.payments_history)
                except:
                    payment_history = []
            
            # Filter payments within selected date range
            for payment in payment_history:
                payment_date_str = payment.get('date', '')
                if payment_date_str:
                    try:
                        if 'T' in payment_date_str:
                            payment_date = datetime.fromisoformat(payment_date_str).date()
                        else:
                            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                        
                        if from_date_obj <= payment_date <= to_date_obj:
                            amount = float(payment.get('amount', 0))
                            method = payment.get('payment_method', 'cash')
                            
                            invoice_rows += f"""
                            <tr>
                                <td class="border">{inv.invoice_no}</td>
                                <td class="border">{payment_date.strftime('%Y-%m-%d')}</td>
                                <td class="border">{inv.customer_name or 'N/A'}</td>
                                <td class="border text-right">{amount:.0f}</td>
                                <td class="border">{method}</td>
                                <td class="border">{payment_date.strftime('%I:%M %p') if 'T' in payment_date_str else '12:00 AM'}</td>
                            </tr>
                            """
                            total_collection += amount
                            invoice_count += 1
                    except:
                        continue
        except:
            continue
    
    # Create HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 15mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 11px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 24px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 15px 0;
                color: #666;
                font-size: 12px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .border {{
                border: 1px solid #000;
                padding: 6px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 8px;
                font-size: 11px;
            }}
            .text-right {{
                text-align: right;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <h1>Customer Invoice Payment Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 13%;">Invoice No</th>
                    <th style="width: 15%;">Payment Date</th>
                    <th style="width: 22%;">Customer Name</th>
                    <th style="width: 15%;">Payment Amount</th>
                    <th style="width: 15%;">Payment Method</th>
                    <th style="width: 15%;">Time</th>
                </tr>
            </thead>
            <tbody>
                {invoice_rows}
                <tr class="total-row">
                    <td class="border" colspan="3" style="text-align: left; font-weight: bold;">TOTAL ({invoice_count} payments)</td>
                    <td class="border text-right" style="font-weight: bold;">{total_collection:.0f}</td>
                    <td class="border" colspan="2"></td>
                </tr>
            </tbody>
        </table>
        
        <div style="margin-top: 20px; border: 2px solid #333; padding: 10px;">
            <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                <span style="font-weight: bold;">Total Collection:</span>
                <span>Rs. {total_collection:.0f}</span>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception as e:
        pdf_content = f"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\ntrailer\n<< /Size 4 /Root 1 0 R >>\n%%EOF"
        encoded_pdf = base64.b64encode(pdf_content.encode()).decode()

    return {"pdf": encoded_pdf}


@router.get("/customized-invoices/excel")
async def get_customized_invoices_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for customer invoices (date-wise payments)
    Returns actual .xlsx file
    """
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query customer invoices
    statement = select(CustomerInvoice).where(
        CustomerInvoice.invoice_no.like("CIN-%")
    )
    result = await db.execute(statement)
    all_invoices = result.scalars().all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Invoice Payments"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    headers = [
        'Invoice No',
        'Payment Date',
        'Customer Name',
        'Payment Amount',
        'Payment Method',
        'Time'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [15, 15, 30, 15, 18, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    total_collection = 0.0
    invoice_count = 0
    row_num = 2
    
    # Write data rows
    for inv in all_invoices:
        try:
            payment_history = []
            if inv.payments_history:
                try:
                    payment_history = json.loads(inv.payments_history)
                except:
                    payment_history = []
            
            for payment in payment_history:
                payment_date_str = payment.get('date', '')
                if payment_date_str:
                    try:
                        if 'T' in payment_date_str:
                            payment_date = datetime.fromisoformat(payment_date_str).date()
                        else:
                            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                        
                        if from_date_obj <= payment_date <= to_date_obj:
                            amount = float(payment.get('amount', 0))
                            method = payment.get('payment_method', 'cash')
                            
                            # Write row data
                            ws.cell(row=row_num, column=1, value=inv.invoice_no).border = thin_border
                            ws.cell(row=row_num, column=2, value=payment_date.strftime('%Y-%m-%d')).border = thin_border
                            ws.cell(row=row_num, column=3, value=inv.customer_name or 'N/A').border = thin_border
                            ws.cell(row=row_num, column=4, value=round(amount, 2)).border = thin_border
                            ws.cell(row=row_num, column=5, value=method).border = thin_border
                            ws.cell(row=row_num, column=6, value=payment_date.strftime('%I:%M %p') if 'T' in payment_date_str else '12:00 AM').border = thin_border
                            
                            # Apply alignments
                            for col in range(1, 7):
                                if col in [4]:  # Amount column
                                    ws.cell(row=row_num, column=col).alignment = right_alignment
                                else:
                                    ws.cell(row=row_num, column=col).alignment = cell_alignment
                            
                            total_collection += amount
                            invoice_count += 1
                            row_num += 1
                    except:
                        continue
        except:
            continue
    
    # Write total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value=f'TOTAL ({invoice_count} payments)').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(total_collection, 2)).font = Font(bold=True)
    
    # Apply border to all cells in total row
    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.border = thin_border
        if col != 4:  # Set empty cells for columns without data
            if col > 4:
                cell.value = ''
    
    # Merge cells for TOTAL label (columns 1-3)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    # Merge empty cells (columns 5-6)
    ws.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=6)
    
    # Add summary section
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value='Total Collection:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=round(total_collection, 2)).font = Font(bold=True)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()
    
    # Encode to base64
    encoded_excel = base64.b64encode(excel_bytes).decode()
    
    return {
        "excel": encoded_excel,
        "filename": f"customer_invoices_{from_date}_to_{to_date}.xlsx"
    }


# ==================== EXPENSE REPORT ENDPOINTS ====================

@router.get("/expenses/pdf")
async def get_expenses_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for expenses (date-wise)
    """
    import base64
    from datetime import datetime
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query expenses for date range
    statement = select(Expense).where(
        and_(
            Expense.expense_date >= from_date_obj,
            Expense.expense_date <= to_date_obj
        )
    )
    result = await db.execute(statement)
    expenses = result.scalars().all()
    
    # Build expense rows for PDF
    expense_rows = ""
    total_amount = 0.0
    
    for exp in expenses:
        expense_rows += f"""
        <tr>
            <td class="border">{exp.expense_date.strftime('%Y-%m-%d')}</td>
            <td class="border">{exp.expense_type}</td>
            <td class="border">{exp.expense}</td>
            <td class="border text-right">{exp.amount:.0f}</td>
            <td class="border">{exp.branch or 'N/A'}</td>
        </tr>
        """
        total_amount += float(exp.amount)
    
    # Create HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 15mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 11px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 24px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 15px 0;
                color: #666;
                font-size: 12px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .border {{
                border: 1px solid #000;
                padding: 6px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 8px;
                font-size: 11px;
            }}
            .text-right {{
                text-align: right;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <h1>Expense Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 15%;">Date</th>
                    <th style="width: 20%;">Expense Type</th>
                    <th style="width: 35%;">Description</th>
                    <th style="width: 15%;">Amount</th>
                    <th style="width: 15%;">Branch</th>
                </tr>
            </thead>
            <tbody>
                {expense_rows}
                <tr class="total-row">
                    <td class="border" colspan="3" style="text-align: left; font-weight: bold;">TOTAL</td>
                    <td class="border text-right" style="font-weight: bold;">{total_amount:.0f}</td>
                    <td class="border"></td>
                </tr>
            </tbody>
        </table>
        
        <div style="margin-top: 20px; border: 2px solid #333; padding: 10px;">
            <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                <span style="font-weight: bold;">Total Expenses:</span>
                <span>Rs. {total_amount:.0f}</span>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception as e:
        pdf_content = f"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\ntrailer\n<< /Size 4 /Root 1 0 R >>\n%%EOF"
        encoded_pdf = base64.b64encode(pdf_content.encode()).decode()

    return {"pdf": encoded_pdf}


@router.get("/expenses/excel")
async def get_expenses_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for expenses (date-wise)
    Returns actual .xlsx file
    """
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query expenses for date range
    statement = select(Expense).where(
        and_(
            Expense.expense_date >= from_date_obj,
            Expense.expense_date <= to_date_obj
        )
    )
    result = await db.execute(statement)
    expenses = result.scalars().all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    headers = [
        'Date',
        'Expense Type',
        'Description',
        'Amount',
        'Branch'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [15, 20, 40, 15, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    total_amount = 0.0
    row_num = 2
    
    # Write data rows
    for exp in expenses:
        ws.cell(row=row_num, column=1, value=exp.expense_date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_num, column=2, value=exp.expense_type).border = thin_border
        ws.cell(row=row_num, column=3, value=exp.expense).border = thin_border
        ws.cell(row=row_num, column=4, value=round(float(exp.amount), 2)).border = thin_border
        ws.cell(row=row_num, column=5, value=exp.branch or 'N/A').border = thin_border
        
        # Apply alignments
        for col in range(1, 6):
            if col in [4]:  # Amount column
                ws.cell(row=row_num, column=col).alignment = right_alignment
            else:
                ws.cell(row=row_num, column=col).alignment = cell_alignment
        
        total_amount += float(exp.amount)
        row_num += 1
    
    # Write total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(total_amount, 2)).font = Font(bold=True)
    
    # Merge cells for TOTAL label
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=5)
    
    # Apply border to total row
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.border = thin_border
        if col > 4:
            cell.value = ''
    
    # Add summary section
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value='Total Expenses:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=round(total_amount, 2)).font = Font(bold=True)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()
    
    # Encode to base64
    encoded_excel = base64.b64encode(excel_bytes).decode()
    
    return {
        "excel": encoded_excel,
        "filename": f"expenses_{from_date}_to_{to_date}.xlsx"
    }


# ==================== STOCK ADJUSTMENT REPORT ENDPOINTS ====================

@router.get("/stock-adjustments/pdf")
async def get_stock_adjustments_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for stock adjustments (date-wise)
    """
    import base64
    from datetime import datetime
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query stock adjustments for date range
    statement = select(StockEntry).where(
        and_(
            StockEntry.type == StockEntryType.ADJUST,
            func.date(StockEntry.created_at) >= from_date_obj,
            func.date(StockEntry.created_at) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    adjustments = result.scalars().all()
    
    # Build adjustment rows for PDF
    adjustment_rows = ""
    total_positive = 0
    total_negative = 0
    
    for adj in adjustments:
        # Get product name
        product_result = await db.execute(select(Product).where(Product.id == adj.product_id))
        product = product_result.scalar_one_or_none()
        product_name = product.name if product else 'Unknown'
        
        # Determine if positive or negative adjustment
        adj_type = "Increase" if adj.qty > 0 else "Decrease"
        abs_qty = abs(adj.qty)
        
        if adj.qty > 0:
            total_positive += adj.qty
        else:
            total_negative += abs(adj.qty)
        
        adjustment_rows += f"""
        <tr>
            <td class="border">{adj.created_at.strftime('%Y-%m-%d')}</td>
            <td class="border">{product_name}</td>
            <td class="border">{adj_type}</td>
            <td class="border text-right">{abs_qty}</td>
            <td class="border">{adj.location or 'N/A'}</td>
            <td class="border">{adj.batch or '-'}</td>
            <td class="border">{adj.created_at.strftime('%I:%M %p')}</td>
        </tr>
        """
    
    # Create HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 15mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 10px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 24px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 15px 0;
                color: #666;
                font-size: 12px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .border {{
                border: 1px solid #000;
                padding: 6px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 8px;
                font-size: 10px;
            }}
            .text-right {{
                text-align: right;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
            }}
            .summary {{
                margin-top: 20px;
                border: 2px solid #333;
                padding: 10px;
            }}
            .summary-row {{
                display: flex;
                justify-content: space-between;
                margin-bottom: 5px;
            }}
        </style>
    </head>
    <body>
        <h1>Stock Adjustment Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 12%;">Date</th>
                    <th style="width: 25%;">Product</th>
                    <th style="width: 12%;">Type</th>
                    <th style="width: 10%;">Qty</th>
                    <th style="width: 15%;">Location</th>
                    <th style="width: 13%;">Batch</th>
                    <th style="width: 13%;">Time</th>
                </tr>
            </thead>
            <tbody>
                {adjustment_rows}
                <tr class="total-row">
                    <td class="border" colspan="3" style="text-align: left; font-weight: bold;">TOTAL</td>
                    <td class="border text-right" style="font-weight: bold;">{total_positive + total_negative}</td>
                    <td class="border" colspan="3"></td>
                </tr>
            </tbody>
        </table>
        
        <div class="summary">
            <div class="summary-row">
                <span style="font-weight: bold;">Stock Increases:</span>
                <span>{total_positive} units</span>
            </div>
            <div class="summary-row">
                <span style="font-weight: bold;">Stock Decreases:</span>
                <span>{total_negative} units</span>
            </div>
            <div class="summary-row">
                <span style="font-weight: bold;">Net Adjustment:</span>
                <span>{total_positive - total_negative} units</span>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception as e:
        pdf_content = f"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\ntrailer\n<< /Size 4 /Root 1 0 R >>\n%%EOF"
        encoded_pdf = base64.b64encode(pdf_content.encode()).decode()

    return {"pdf": encoded_pdf}


@router.get("/stock-adjustments/excel")
async def get_stock_adjustments_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for stock adjustments (date-wise)
    Returns actual .xlsx file
    """
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Query stock adjustments for date range
    statement = select(StockEntry).where(
        and_(
            StockEntry.type == StockEntryType.ADJUST,
            func.date(StockEntry.created_at) >= from_date_obj,
            func.date(StockEntry.created_at) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    adjustments = result.scalars().all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    headers = [
        'Date',
        'Product Name',
        'Adjustment Type',
        'Quantity',
        'Location',
        'Batch',
        'Time'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [15, 30, 15, 12, 20, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    total_positive = 0
    total_negative = 0
    row_num = 2
    
    # Write data rows
    for adj in adjustments:
        # Get product name
        try:
            product_result = await db.execute(select(Product).where(Product.id == adj.product_id))
            product = product_result.scalar_one_or_none()
            product_name = product.name if product else 'Unknown'
        except:
            product_name = 'Unknown'
        
        # Determine adjustment type
        adj_type = "Increase" if adj.qty > 0 else "Decrease"
        abs_qty = abs(adj.qty)
        
        if adj.qty > 0:
            total_positive += adj.qty
        else:
            total_negative += abs(adj.qty)
        
        # Write row data
        ws.cell(row=row_num, column=1, value=adj.created_at.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_num, column=2, value=product_name).border = thin_border
        ws.cell(row=row_num, column=3, value=adj_type).border = thin_border
        ws.cell(row=row_num, column=4, value=abs_qty).border = thin_border
        ws.cell(row=row_num, column=5, value=adj.location or 'N/A').border = thin_border
        ws.cell(row=row_num, column=6, value=adj.batch or '-').border = thin_border
        ws.cell(row=row_num, column=7, value=adj.created_at.strftime('%I:%M %p')).border = thin_border
        
        # Apply alignments
        for col in range(1, 8):
            if col in [4]:  # Quantity column
                ws.cell(row=row_num, column=col).alignment = right_alignment
            else:
                ws.cell(row=row_num, column=col).alignment = cell_alignment
        
        row_num += 1
    
    # Write total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_positive + total_negative).font = Font(bold=True)
    
    # Merge cells for TOTAL label (columns 1-3)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    # Merge empty cells (columns 5-7)
    ws.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=7)
    
    # Apply border to total row (only non-merged cells)
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.border = thin_border
    
    # Add summary section
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value='Stock Increases:').font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=total_positive)
    
    ws.cell(row=summary_row + 1, column=1, value='Stock Decreases:').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=total_negative)
    
    ws.cell(row=summary_row + 2, column=1, value='Net Adjustment:').font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=total_positive - total_negative)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()
    
    # Encode to base64
    encoded_excel = base64.b64encode(excel_bytes).decode()
    
    return {
        "excel": encoded_excel,
        "filename": f"stock_adjustments_{from_date}_to_{to_date}.xlsx"
    }


# ==================== REFUND REPORT ENDPOINTS ====================

@router.get("/refunds/pdf")
async def get_refunds_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for refunds in date range
    """
    import base64
    from datetime import datetime

    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Query refunds for walk-in invoices only (SIN- prefix)
    from ..models.refund import Refund
    # Convert to datetime for proper comparison
    from_date_datetime = datetime.combine(from_date_obj, datetime.min.time())
    to_date_datetime = datetime.combine(to_date_obj, datetime.max.time())
    
    # Join refunds with invoices and filter by SIN- prefix
    refunds_statement = select(Refund).join(
        Invoice, Refund.invoice_id == Invoice.id
    ).where(
        and_(
            Refund.created_at >= from_date_datetime,
            Refund.created_at <= to_date_datetime,
            Invoice.invoice_no.like('SIN-%')  # Only walk-in invoices
        )
    ).order_by(Refund.created_at.desc())
    
    refunds_result = await db.execute(refunds_statement)
    refunds = refunds_result.scalars().all()

    # Build refund rows for PDF
    refund_rows = ""
    total_refund_amount = 0.0

    for refund in refunds:
        try:
            # Get invoice number (already filtered to SIN- prefix)
            invoice_result = await db.execute(select(Invoice).where(Invoice.id == refund.invoice_id))
            invoice = invoice_result.scalar_one_or_none()
            invoice_no = invoice.invoice_no if invoice else "N/A"
            
            # Parse refund items to get product details
            refund_items = json.loads(refund.items) if refund.items else []
            
            for item in refund_items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity_returned', 0)
                unit_price = item.get('unit_price', 0)
                item_total = quantity * unit_price
                item_cost = item_total * 0.7  # 70% cost calculation
                
                refund_amount = float(refund.amount)
                total_refund_amount += refund_amount
                
                refund_rows += f"""
                <tr>
                    <td class="border">{invoice_no}</td>
                    <td class="border">{refund.created_at.strftime('%d-%m-%Y') if refund.created_at else ''}</td>
                    <td class="border">{product_name}</td>
                    <td class="border text-right">{quantity}</td>
                    <td class="border text-right">{unit_price:.0f}</td>
                    <td class="border text-right">{item_total:.0f}</td>
                    <td class="border text-right">{item_cost:.0f}</td>
                </tr>
                """
        except:
            continue

    # Create HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 15mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 14px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 28px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 15px 0;
                color: #666;
                font-size: 15px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            .border {{
                border: 1px solid #000;
                padding: 10px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 12px;
                font-size: 14px;
            }}
            .text-right {{
                text-align: right;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
                font-size: 15px;
            }}
        </style>
    </head>
    <body>
        <h1>Refund Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>

        <table>
            <thead>
                <tr>
                    <th style="width: 15%;">Invoice No</th>
                    <th style="width: 12%;">Date</th>
                    <th style="width: 28%;">Product</th>
                    <th style="width: 8%;">Quantity</th>
                    <th style="width: 10%;">Price</th>
                    <th style="width: 12%;">Amount</th>
                    <th style="width: 12%;">Cost</th>
                </tr>
            </thead>
            <tbody>
                {refund_rows}
                <tr class="total-row">
                    <td class="border" colspan="5" style="text-align: left; font-weight: bold;">TOTAL REFUND</td>
                    <td class="border text-right" style="font-weight: bold;">{total_refund_amount:.0f}</td>
                    <td class="border"></td>
                </tr>
            </tbody>
        </table>
    </body>
    </html>
    """

    # Generate PDF using weasyprint (same as customer_invoice.py)
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception as e:
        # Fallback - minimal PDF
        encoded_pdf = base64.b64encode(b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\ntrailer\n<< /Size 4 /Root 1 0 R >>\n%%EOF").decode()

    return {
        "pdf": encoded_pdf,
        "filename": f"refund_report_{from_date}_to_{to_date}.pdf"
    }


@router.get("/refunds/excel")
async def get_refunds_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    branch: str = Query("European Sports Light House"),
    current_user: User = Depends(admin_employee_required_from_session()),  # Admin and Employee only (NOT Cashier)
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for refunds in date range
    """
    import base64
    import io
    from datetime import datetime
    from ..models.refund import Refund
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Query refunds for walk-in invoices only (SIN- prefix)
    from ..models.refund import Refund
    # Convert to datetime for proper comparison
    from_date_datetime = datetime.combine(from_date_obj, datetime.min.time())
    to_date_datetime = datetime.combine(to_date_obj, datetime.max.time())
    
    # Join refunds with invoices and filter by SIN- prefix
    refunds_statement = select(Refund).join(
        Invoice, Refund.invoice_id == Invoice.id
    ).where(
        and_(
            Refund.created_at >= from_date_datetime,
            Refund.created_at <= to_date_datetime,
            Invoice.invoice_no.like('SIN-%')  # Only walk-in invoices
        )
    ).order_by(Refund.created_at.desc())
    
    refunds_result = await db.execute(refunds_statement)
    refunds = refunds_result.scalars().all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Refund Report"

    # Define styles (matching other reports)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header
    headers = [
        'Invoice No',
        'Date',
        'Product Name',
        'Quantity',
        'Unit Price',
        'Amount',
        'Cost'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    column_widths = [15, 12, 30, 10, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    total_refund_amount = 0.0
    total_cost = 0.0
    row_num = 2

    # Write data rows
    for refund in refunds:
        try:
            # Get invoice number (already filtered to SIN- prefix)
            invoice_result = await db.execute(select(Invoice).where(Invoice.id == refund.invoice_id))
            invoice = invoice_result.scalar_one_or_none()
            invoice_no = invoice.invoice_no if invoice else "N/A"
            
            # Parse refund items
            refund_items = json.loads(refund.items) if refund.items else []
            
            for item in refund_items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity_returned', 0)
                unit_price = item.get('unit_price', 0)
                item_total = quantity * unit_price
                item_cost = item_total * 0.7  # 70% cost calculation
                
                refund_amount = float(refund.amount)
                total_refund_amount += refund_amount
                total_cost += item_cost

                # Write row data
                ws.cell(row=row_num, column=1, value=invoice_no).border = thin_border
                ws.cell(row=row_num, column=2, value=refund.created_at.strftime('%d-%m-%Y') if refund.created_at else '').border = thin_border
                ws.cell(row=row_num, column=3, value=product_name).border = thin_border
                ws.cell(row=row_num, column=4, value=quantity).border = thin_border
                ws.cell(row=row_num, column=5, value=round(unit_price, 2)).border = thin_border
                ws.cell(row=row_num, column=6, value=round(item_total, 2)).border = thin_border
                ws.cell(row=row_num, column=7, value=round(item_cost, 2)).border = thin_border

                # Apply alignments
                for col in range(1, 8):
                    if col == 4:  # Quantity
                        ws.cell(row=row_num, column=col).alignment = cell_alignment
                    elif col in [5, 6, 7]:  # Numeric columns
                        ws.cell(row=row_num, column=col).alignment = right_alignment
                    else:
                        ws.cell(row=row_num, column=col).alignment = cell_alignment

                row_num += 1
        except:
            continue

    # Write total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value='TOTAL REFUND').font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total_refund_amount, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(total_cost, 2)).font = Font(bold=True)

    # Merge cells for TOTAL label
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)

    # Style total row
    total_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.border = thin_border
        if col <= 7:
            cell.fill = total_fill
        cell.alignment = cell_alignment

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()

    # Encode to base64
    encoded_excel = base64.b64encode(excel_bytes).decode()

    return {
        "excel": encoded_excel,
        "filename": f"refund_report_{from_date}_to_{to_date}.xlsx"
    }
