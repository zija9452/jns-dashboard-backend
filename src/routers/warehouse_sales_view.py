from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Optional, List, Dict, Any
import json
from calendar import monthrange
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from weasyprint import HTML

from src.database import get_db
from src.models.user import User
from src.models.warehouse_invoice import WarehouseInvoice
from src.models.stock_entry import StockEntry, StockEntryType
from src.models.product import Product
from src.auth.session_auth import admin_cashier_employee_required_from_session

router = APIRouter(prefix="/warehouse-salesview", tags=["warehouse-salesview"])

@router.get("/dashboard/stats")
async def get_warehouse_dashboard_stats(
    from_date: str = Query(None, description="Start date in YYYY-MM-DD format"),
    to_date: str = Query(None, description="End date in YYYY-MM-DD format"),
    month: int = None,
    year: int = None,
    current_user: User = Depends(admin_cashier_employee_required_from_session()),
    db: AsyncSession = Depends(get_db)
):
    """
    Get warehouse dashboard statistics for a specific date range or month.
    """
    today = datetime.now().date()
    user_role = current_user.role.name if current_user and current_user.role else "admin"

    # Set date range
    if from_date and to_date:
        try:
            first_day = datetime.fromisoformat(from_date).date()
            data_end_date = datetime.fromisoformat(to_date).date()
            if first_day > data_end_date:
                raise HTTPException(status_code=400, detail="from_date must be before or equal to to_date")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        if month is None:
            month = today.month
        if year is None:
            year = today.year
        if month < 1 or month > 12:
            raise HTTPException(status_code=400, detail="Invalid month. Must be 1-12")
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])
        data_end_date = min(last_day, today)

    # Try cache
    from ..utils.cache import cache
    cache_params = {
        "from_date": str(first_day),
        "to_date": str(data_end_date),
        "user_role": user_role,
        "type": "warehouse_dashboard"
    }
    cached_result = await cache.get_dashboard_stats(cache_params)
    if cached_result:
        return cached_result

    # ==================== WAREHOUSE SALES CALCULATION ====================
    sales_statement = select(func.sum(WarehouseInvoice.amount_paid)).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= first_day,
            func.date(WarehouseInvoice.payment_date) <= data_end_date
        )
    )
    sales_result = await db.execute(sales_statement)
    total_sales = float(sales_result.scalar_one_or_none() or 0)

    # ==================== WAREHOUSE PURCHASE CALCULATION ====================
    purchase_statement = select(func.sum(StockEntry.qty * StockEntry.cost_price)).where(
        and_(
            StockEntry.type == StockEntryType.IN,
            StockEntry.location == "warehouse",
            func.date(StockEntry.created_at) >= first_day,
            func.date(StockEntry.created_at) <= data_end_date
        )
    )
    purchase_result = await db.execute(purchase_statement)
    total_purchase = float(purchase_result.scalar_one_or_none() or 0)

    # ==================== STOCK DATA ====================
    stock_result = await db.execute(
        select(
            func.count(Product.id).filter(
                and_(
                    Product.is_warehouse_product == True,
                    Product.warehouse_stock < Product.warehouse_limited_qty
                )
            ).label('short_stock'),
            func.count(Product.id).filter(
                Product.is_warehouse_product == True
            ).label('total_products'),
            func.count(Product.id).filter(
                and_(
                    Product.is_warehouse_product == True,
                    Product.stock_level < Product.limited_qty
                )
            ).label('short_shop_stock'),
            func.count(Product.id).filter(
                and_(
                    Product.is_warehouse_product == True,
                    (Product.limited_qty - Product.stock_level) > Product.warehouse_stock
                )
            ).label('urgent_buy')
        )
    )
    stock_row = stock_result.one()
    short_stock = int(stock_row.short_stock or 0)
    total_products = int(stock_row.total_products or 0)
    short_shop_stock = int(stock_row.short_shop_stock or 0)
    urgent_buy = int(stock_row.urgent_buy or 0)

    # ==================== DAILY CHART DATA ====================
    daily_sales_statement = select(
        func.date(WarehouseInvoice.payment_date).label('date'),
        func.sum(WarehouseInvoice.amount_paid).label('total')
    ).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= first_day,
            func.date(WarehouseInvoice.payment_date) <= data_end_date
        )
    ).group_by(func.date(WarehouseInvoice.payment_date))

    daily_sales_result = await db.execute(daily_sales_statement)
    daily_sales = {row.date: float(row.total) for row in daily_sales_result.all()}

    daily_purchase_statement = select(
        func.date(StockEntry.created_at).label('date'),
        func.sum(StockEntry.qty * StockEntry.cost_price).label('total')
    ).where(
        and_(
            StockEntry.type == StockEntryType.IN,
            StockEntry.location == "warehouse",
            func.date(StockEntry.created_at) >= first_day,
            func.date(StockEntry.created_at) <= data_end_date
        )
    ).group_by(func.date(StockEntry.created_at))

    daily_purchase_result = await db.execute(daily_purchase_statement)
    daily_purchases = {row.date: float(row.total) for row in daily_purchase_result.all()}

    delta = timedelta(days=1)
    current_date = first_day
    dates = []
    sales_data = []
    purchases_data = []
    
    while current_date <= data_end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        dates.append(date_str)
        sales_data.append(round(daily_sales.get(current_date, 0.0), 2))
        purchases_data.append(round(daily_purchases.get(current_date, 0.0), 2))
        current_date += delta

    response = {
        "totalSales": round(total_sales, 2),
        "totalExpense": 0.0,
        "totalPurchase": round(total_purchase, 2),
        "outOfStock": 0,
        "shortStock": short_stock,
        "totalProducts": total_products,
        "shortShopStock": short_shop_stock,
        "urgentBuy": urgent_buy,
        "adminUser": current_user.username if current_user else "Admin",
        "userRole": user_role,
        "openingBalance": 0.0,
        "chartData": {
            "dates": dates,
            "sales": sales_data,
            "purchases": purchases_data
        },
        "dateRange": {
            "from": first_day.strftime('%Y-%m-%d'),
            "to": data_end_date.strftime('%Y-%m-%d')
        }
    }

    await cache.set_dashboard_stats(cache_params, response, ttl=300)
    return response

@router.get("/sales-view")
async def get_warehouse_sales_view(
    from_date: str = Query(...),
    to_date: str = Query(...),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    )

    if search:
        statement = statement.where(
            or_(
                WarehouseInvoice.invoice_no.ilike(f"%{search}%"),
                WarehouseInvoice.customer_name.ilike(f"%{search}%")
            )
        )

    statement = statement.order_by(WarehouseInvoice.created_at.asc())
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    invoice_list = []
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            total_invoice_paid = float(inv.amount_paid)
            total_invoice_discount = float(inv.discounts)
            
            first_item = True
            for item in items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                item_total = quantity * unit_price - item_discount
                item_cost = float(item.get('cost_price', 0)) * quantity
                
                amount_paid = total_invoice_paid if first_item else 0.0
                
                invoice_list.append({
                    "id": str(inv.id),
                    "invoice_no": inv.invoice_no,
                    "product_name": product_name,
                    "total_amount": float(item_total),
                    "amount_paid": amount_paid,
                    "balance_due": 0.0,
                    "payment_status": inv.payment_status,
                    "payment_method": inv.payment_method or "cash",
                    "quantity": quantity,
                    "discount": item_discount if first_item else 0.0,
                    "total_discount": total_invoice_discount if first_item else 0.0,
                    "cost": item_cost,
                    "created_at": inv.payment_date.isoformat() if inv.payment_date else inv.created_at.isoformat(),
                    "customer_name": inv.customer_name
                })
                first_item = False
        except Exception:
            continue

    return {"invoices": invoice_list, "total": len(invoice_list)}

@router.get("/summary")
async def get_warehouse_summary(
    from_date: str = Query(...),
    to_date: str = Query(...),
    current_user: User = Depends(admin_cashier_employee_required_from_session()),
    db: AsyncSession = Depends(get_db)
):
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    total_sale = 0.0
    total_cost = 0.0
    total_discount = 0.0
    
    for inv in invoices:
        total_sale += float(inv.amount_paid)
        total_discount += float(inv.discounts)
        try:
            items = json.loads(inv.items) if inv.items else []
            for item in items:
                total_cost += float(item.get('cost_price', 0)) * item.get('quantity', 0)
        except:
            continue

    net_profit = total_sale - total_cost
    net_cash = total_sale

    return {
        "opening": 0,
        "totalSale": total_sale,
        "grossProfit": total_sale - total_cost,
        "totalExpense": 0.0,
        "totalRecovery": 0,
        "vendorPayments": 0,
        "netCash": net_cash,
        "totalPurchase": 0,
        "totalRefund": 0,
        "netProfit": net_profit,
        "warehouse_sales": total_sale
    }

@router.get("/sales-view/pdf")
async def get_warehouse_sales_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    db: AsyncSession = Depends(get_db)
):
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    ).order_by(WarehouseInvoice.payment_date.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()

    invoice_rows = ""
    total_amount_paid = 0.0
    
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            total_invoice_paid = float(inv.amount_paid)
            first_item = True
            for item in items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                unit_cost = float(item.get('cost_price', 0))
                row_total = (quantity * unit_price) - item_discount
                amount_paid = total_invoice_paid if first_item else 0.0
                total_amount_paid += amount_paid

                invoice_rows += f"""
                <tr>
                    <td class="border">{inv.invoice_no}</td>
                    <td class="border">{product_name}</td>
                    <td class="border">{inv.customer_name or 'N/A'}</td>
                    <td class="border text-right">{unit_cost:.0f}</td>
                    <td class="border text-right">{unit_price:.0f}</td>
                    <td class="border text-center">{quantity}</td>
                    <td class="border text-right">{row_total:.0f}</td>
                    <td class="border text-right">{item_discount if first_item else 0}</td>
                    <td class="border text-right">{amount_paid:.0f}</td>
                    <td class="border">{inv.payment_date.strftime('%I:%M:%S %p') if inv.payment_date else inv.created_at.strftime('%I:%M:%S %p')}</td>
                </tr>
                """
                first_item = False
        except Exception:
            continue

    current_date = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 10mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10px; }}
            h1 {{ text-align: center; font-size: 20px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            .border {{ border: 1px solid #000; padding: 4px; }}
            th {{ background-color: #444; color: white; padding: 6px; }}
            .text-right {{ text-align: right; }}
            .total-row {{ background-color: #f0f0f0; font-weight: bold; }}
        </style>
    </head>
    <body>
        <h1>Warehouse Sales Report</h1>
        <p style="text-align: center;">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        <table>
            <thead>
                <tr>
                    <th>Order ID</th><th>Product</th><th>Customer</th><th>Cost</th><th>Price</th><th>Qty</th><th>Total</th><th>Discount</th><th>Amount Paid</th><th>Time</th>
                </tr>
            </thead>
            <tbody>
                {invoice_rows}
                <tr class="total-row">
                    <td class="border" colspan="8">GRAND TOTAL</td><td class="border text-right">{total_amount_paid:.0f}</td><td class="border"></td>
                </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    pdf_bytes = HTML(string=html_content).write_pdf()
    return {"pdf": base64.b64encode(pdf_bytes).decode()}

@router.get("/sales-view/excel")
async def get_warehouse_sales_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    db: AsyncSession = Depends(get_db)
):
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    ).order_by(WarehouseInvoice.payment_date.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Sales"

    headers = ['Order ID', 'Product', 'Customer', 'Cost', 'Price', 'Qty', 'Total', 'Discount', 'Amount Paid', 'Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    row_num = 2
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            total_invoice_paid = float(inv.amount_paid)
            first_item = True
            for item in items:
                ws.cell(row=row_num, column=1, value=inv.invoice_no)
                ws.cell(row=row_num, column=2, value=item.get('product_name'))
                ws.cell(row=row_num, column=3, value=inv.customer_name)
                ws.cell(row=row_num, column=4, value=float(item.get('cost_price', 0)))
                ws.cell(row=row_num, column=5, value=float(item.get('unit_price', 0)))
                ws.cell(row=row_num, column=6, value=item.get('quantity'))
                ws.cell(row=row_num, column=7, value=(item.get('quantity', 0) * float(item.get('unit_price', 0))) - float(item.get('discount', 0)))
                ws.cell(row=row_num, column=8, value=float(item.get('discount', 0)) if first_item else 0)
                ws.cell(row=row_num, column=9, value=total_invoice_paid if first_item else 0)
                ws.cell(row=row_num, column=10, value=inv.payment_date.strftime('%H:%M:%S') if inv.payment_date else inv.created_at.strftime('%H:%M:%S'))
                row_num += 1
                first_item = False
        except Exception:
            continue

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return {"excel": base64.b64encode(output.read()).decode(), "filename": f"warehouse_sales_{from_date}_to_{to_date}.xlsx"}
