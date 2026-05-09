from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Dict, Optional, Any
import uuid
import json
import base64
import io
from datetime import datetime, date, timedelta
from decimal import Decimal
from sqlalchemy import select, func, or_, and_
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from weasyprint import HTML

from ..database.database import get_db
from ..models.warehouse_invoice import WarehouseInvoice
from ..models.product import Product
from ..models.warehouse_customer import WarehouseCustomer
from ..models.salesman import Salesman
from ..models.user import User
from ..models.stock_entry import StockEntry, StockEntryType
from ..auth.session_auth import admin_cashier_employee_required_from_session, admin_employee_required_from_session

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/warehouse-invoice", tags=["warehouse-invoice"])

@router.post("/create")
async def create_warehouse_invoice(
    request_data: dict,
    current_user: User = Depends(admin_cashier_employee_required_from_session()),
    db: AsyncSession = Depends(get_db)
):
    """
    Create a new warehouse invoice and save to warehouse_invoices table.
    Decreases warehouse_stock and increases stock_level (shop stock).
    """
    # Extract data from request body
    order_items = request_data.get('items', [])
    customer_id = request_data.get('customer_id')
    payment_method = request_data.get('payment_method', 'cash')
    payment_date_str = request_data.get('payment_date', datetime.now().isoformat())
    
    selected_date = datetime.fromisoformat(payment_date_str).date()
    current_time = datetime.now().time()
    payment_date = datetime.combine(selected_date, current_time)

    manual_discount = float(request_data.get('manual_discount', 0))
    # Get amount_paid from request, if not provided it will be calculated later
    requested_amount_paid = request_data.get('amount_paid')
    notes = request_data.get('notes', '')

    if not order_items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Order items are required"
        )

    # Validate warehouse customer
    customer_name = "European Sports"
    customer_id_uuid = None
    customer_obj = None
    if customer_id:
        try:
            customer_id_uuid = uuid.UUID(str(customer_id))
            customer_result = await db.execute(select(WarehouseCustomer).where(WarehouseCustomer.id == customer_id_uuid))
            customer_obj = customer_result.scalar_one_or_none()
            if customer_obj:
                customer_name = customer_obj.name
            else:
                logger.error(f"Warehouse customer with ID {customer_id} not found")
        except ValueError:
            logger.error(f"Invalid UUID format for customer_id: {customer_id}")

    # Process each order item and update inventory (Transfer logic)
    items_list = []
    total_amount = Decimal('0')
    total_discount = 0.0

    try:
        for item in order_items:
            pro_name = item.get('pro_name')
            quantity = int(item.get('pro_quantity', 0))
            unit_price = float(item.get('unit_price', 0))
            discount = float(item.get('discount', 0))

            # Verify product exists
            product_result = await db.execute(select(Product).where(Product.name == pro_name))
            product = product_result.scalar_one_or_none()
            if not product:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Product '{pro_name}' not found"
                )

            # Check warehouse stock
            if (product.warehouse_stock or 0) < quantity:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Insufficient warehouse stock for '{pro_name}'. Available: {product.warehouse_stock or 0}"
                )

            # TRANSFER LOGIC: Warehouse -> Shop
            product.warehouse_stock = (product.warehouse_stock or 0) - quantity
            product.stock_level = (product.stock_level or 0) + quantity
            db.add(product)

            # Stock Entries
            stock_out = StockEntry(
                product_id=product.id,
                qty=-quantity,
                type=StockEntryType.OUT,
                location="warehouse",
                ref=f"W_INV_TRANSFER_{datetime.now().strftime('%Y%m%d')}"
            )
            db.add(stock_out)

            stock_in = StockEntry(
                product_id=product.id,
                qty=quantity,
                type=StockEntryType.IN,
                location="shop",
                ref=f"W_INV_TRANSFER_{datetime.now().strftime('%Y%m%d')}"
            )
            db.add(stock_in)

            item_total_before_discount = quantity * unit_price
            item_obj = {
                "product_name": str(pro_name),
                "product_id": str(product.id),
                "quantity": quantity,
                "unit_price": unit_price,
                "total_price": item_total_before_discount,
                "discount": discount,
                "cat_name": str(item.get('cat_name', '')),
                "cost_price": float(product.warehouse_cost) if product.warehouse_cost else 0.0,
            }
            items_list.append(item_obj)
            total_amount += Decimal(str(item_total_before_discount))
            total_discount += discount

        total_discount += manual_discount
        net_total = total_amount - Decimal(str(total_discount))
        
        # Use requested amount paid or default to net_total
        if requested_amount_paid is not None:
            amount_paid = Decimal(str(requested_amount_paid))
        else:
            amount_paid = net_total
            
        balance_due = net_total - amount_paid
        
        # Update customer balance if there is a balance due
        if customer_obj and balance_due != 0:
            customer_obj.cus_balance = (customer_obj.cus_balance or Decimal('0')) + balance_due
            db.add(customer_obj)

        # Determine payment status
        payment_status = "paid"
        if balance_due > 0:
            payment_status = "partial" if amount_paid > 0 else "unpaid"
        elif balance_due < 0:
            payment_status = "overpaid"

        # Generate unique invoice number (WIN prefix)
        await db.execute(select(func.pg_advisory_lock(123461)))
        try:
            prefix_pattern = "WIN-%"
            statement = select(func.max(WarehouseInvoice.invoice_no)).where(
                WarehouseInvoice.invoice_no.like(prefix_pattern)
            )
            result = await db.execute(statement)
            max_invoice_no = result.scalar_one_or_none()

            if max_invoice_no:
                try:
                    parts = max_invoice_no.split("-")
                    next_seq = int(parts[-1]) + 1
                    seq_number = f"{next_seq:04d}"
                except:
                    seq_number = "0001"
            else:
                seq_number = "0001"
            
            invoice_no = f"WIN-{seq_number}"

            # Create WarehouseInvoice
            invoice_obj = WarehouseInvoice(
                id=uuid.uuid4(),
                invoice_no=invoice_no,
                customer_id=customer_id_uuid,
                customer_name=customer_name,
                items=json.dumps(items_list),
                totals=json.dumps({
                    "subtotal": float(total_amount),
                    "tax": 0.0,
                    "discount": float(total_discount),
                    "total": float(net_total),
                    "amount_paid": float(amount_paid),
                    "balance_due": float(balance_due),
                    "payment_status": payment_status
                }),
                total_amount=net_total,
                amount_paid=amount_paid,
                payment_status=payment_status,
                payments_history=json.dumps([{
                    "amount": float(amount_paid),
                    "payment_method": payment_method,
                    "date": payment_date.isoformat(),
                    "description": "Warehouse Transfer Payment"
                }]),
                discounts=Decimal(str(total_discount)),
                payment_method=payment_method,
                payment_date=payment_date,
                notes=notes,
                created_by=current_user.id
            )
            
            db.add(invoice_obj)
            await db.commit()
            await db.refresh(invoice_obj)

            # Generate PDF receipt
            from .walkin_invoice import generate_walkin_receipt_pdf
            pdf_data = generate_walkin_receipt_pdf(
                invoice_no=invoice_no,
                customer_name=customer_name,
                team_name="",
                items=items_list,
                total_amount=float(total_amount),
                total_discount=float(total_discount),
                amount_paid=float(amount_paid),
                balance_due=float(balance_due),
                payment_method=payment_method,
                payment_status=payment_status,
                created_at=payment_date,
                bill_type="WAREHOUSE RECEIPT"
            )
            
            return {"pdf": pdf_data, "invoice_no": invoice_no, "invoice_id": str(invoice_obj.id)}
        finally:
            await db.execute(select(func.pg_advisory_unlock(123461)))
    except Exception as e:
        logger.error(f"Error creating warehouse invoice: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Error creating warehouse invoice: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@router.get("/products")
async def get_warehouse_products(
    db: AsyncSession = Depends(get_db)
):
    """
    Get products that belong to warehouse and have stock
    """
    statement = select(Product).where(
        Product.is_warehouse_product == True,
        Product.warehouse_stock > 0
    )
    result = await db.execute(statement)
    products = result.scalars().all()
    
    return [
        {
            "id": str(p.id),
            "product_name": p.name,
            "category": p.category,
            "price": float(p.cost_price) if p.cost_price else 0.0,
            "stock": p.warehouse_stock,
            "barcode": p.barcode
        }
        for p in products
    ]

@router.get("/sales-view")
async def get_warehouse_sales_view(
    from_date: str = Query(...),
    to_date: str = Query(...),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Get warehouse invoices for sales view with date filtering.
    Returns item-wise rows to match the main sales view UI.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    )

    if search:
        statement = statement.where(
            or_(
                WarehouseInvoice.invoice_no.ilike(f"%{search}%"),
                WarehouseInvoice.customer_name.ilike(f"%{search}%")
            )
        )

    statement = statement.order_by(WarehouseInvoice.created_at.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    invoice_list = []
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            totals = json.loads(inv.totals) if inv.totals else {}
            
            total_invoice_paid = float(inv.amount_paid)
            total_invoice_discount = float(inv.discounts)
            
            first_item = True
            for item in items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                item_total = quantity * unit_price - item_discount
                item_cost = float(item.get('cost_price', 0)) * quantity
                
                amount_paid = total_invoice_paid if first_item else 0.0
                item_discount_display = item_discount if first_item else 0.0
                total_discount_display = total_invoice_discount if first_item else 0.0
                
                invoice_list.append({
                    "id": str(inv.id),
                    "invoice_no": inv.invoice_no,
                    "product_name": product_name,
                    "total_amount": float(item_total),
                    "amount_paid": amount_paid,
                    "balance_due": 0.0,
                    "payment_status": inv.payment_status,
                    "payment_method": inv.payment_method or "cash",
                    "quantity": quantity,
                    "discount": item_discount_display,
                    "total_discount": total_discount_display,
                    "cost": item_cost,
                    "created_at": inv.payment_date.isoformat() if inv.payment_date else inv.created_at.isoformat(),
                    "customer_name": inv.customer_name
                })
                first_item = False
        except Exception as e:
            logger.error(f"Error processing warehouse invoice {inv.invoice_no}: {str(e)}")
            continue

    return {"invoices": invoice_list, "total": len(invoice_list)}

@router.get("/summary")
async def get_warehouse_summary(
    from_date: str = Query(...),
    to_date: str = Query(...),
    current_user: User = Depends(admin_cashier_employee_required_from_session()),
    db: AsyncSession = Depends(get_db)
):
    """
    Get warehouse sales summary for the given date range.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    # Calculate warehouse sales
    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    total_sale = 0.0
    total_cost = 0.0
    total_discount = 0.0
    
    for inv in invoices:
        total_sale += float(inv.amount_paid)
        total_discount += float(inv.discounts)
        try:
            items = json.loads(inv.items) if inv.items else []
            for item in items:
                total_cost += float(item.get('cost_price', 0)) * item.get('quantity', 0)
        except:
            continue

    # Get expenses for warehouse (if any)
    # For now, we'll keep it simple as warehouse specific expenses might not be tracked separately yet
    total_expense = 0.0 
    
    net_profit = total_sale - total_cost - total_expense
    net_cash = total_sale - total_expense

    return {
        "opening": 0, # Warehouse might not have daily cash opening like shop
        "totalSale": total_sale,
        "grossProfit": total_sale - total_cost,
        "totalExpense": total_expense,
        "totalRecovery": 0,
        "vendorPayments": 0,
        "netCash": net_cash,
        "totalPurchase": 0, # Can be calculated from StockEntry IN for warehouse
        "totalRefund": 0,
        "netProfit": net_profit,
        "warehouse_sales": total_sale
    }

@router.get("/sales-view/pdf")
async def get_warehouse_sales_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for warehouse sales.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    ).order_by(WarehouseInvoice.payment_date.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()

    invoice_rows = ""
    total_amount_paid = 0.0
    total_cost = 0.0
    
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            total_invoice_paid = float(inv.amount_paid)
            
            first_item = True
            for item in items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                
                # Logic:
                # Cost = unit cost
                # Total = (price * quantity) - discount
                # Amount Paid = total_invoice_paid on first item, else 0
                unit_cost = float(item.get('cost_price', 0))
                row_total = (quantity * unit_price) - item_discount
                amount_paid = total_invoice_paid if first_item else 0.0
                
                total_amount_paid += amount_paid
                total_cost += (unit_cost * quantity)

                invoice_rows += f"""
                <tr>
                    <td class="border">{inv.invoice_no}</td>
                    <td class="border">{product_name}</td>
                    <td class="border">{inv.customer_name or 'N/A'}</td>
                    <td class="border text-right">{unit_cost:.0f}</td>
                    <td class="border text-right">{unit_price:.0f}</td>
                    <td class="border text-center">{quantity}</td>
                    <td class="border text-right">{row_total:.0f}</td>
                    <td class="border text-right">{item_discount if first_item else 0}</td>
                    <td class="border text-right">{amount_paid:.0f}</td>
                    <td class="border">{inv.payment_date.strftime('%I:%M:%S %p') if inv.payment_date else inv.created_at.strftime('%I:%M:%S %p')}</td>
                </tr>
                """
                first_item = False
        except Exception as e:
            logger.error(f"Error parsing invoice {inv.invoice_no} for PDF: {str(e)}")
            continue

    current_date = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 10mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 10px;
                margin: 0;
                padding: 0;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 5px 0;
                font-size: 20px;
                font-weight: bold;
            }}
            .date-range {{
                text-align: center;
                margin: 0 0 10px 0;
                color: #666;
                font-size: 10px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
            }}
            .border {{
                border: 1px solid #000;
                padding: 4px;
            }}
            th {{
                background-color: #444;
                color: white;
                font-weight: bold;
                text-align: left;
                padding: 6px;
                font-size: 10px;
            }}
            .text-right {{
                text-align: right;
            }}
            .text-center {{
                text-align: center;
            }}
            .total-row {{
                background-color: #f0f0f0;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <h1>Warehouse Sales Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 8%;">Order ID</th>
                    <th style="width: 15%;">Product</th>
                    <th style="width: 12%;">Customer</th>
                    <th style="width: 8%;">Cost</th>
                    <th style="width: 8%;">Price</th>
                    <th style="width: 5%;">Qty</th>
                    <th style="width: 8%;">Total</th>
                    <th style="width: 8%;">Discount</th>
                    <th style="width: 8%;">Amount Paid</th>
                    <th style="width: 15%;">Time</th>
                </tr>
            </thead>
            <tbody>
                {invoice_rows}
                <tr class="total-row">
                    <td class="border" colspan="8" style="text-align: left;">GRAND TOTAL</td>
                    <td class="border text-right">{total_amount_paid:.0f}</td>
                    <td class="border"></td>
                </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    try:
        pdf_bytes = HTML(string=html_content).write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
        return {"pdf": encoded_pdf}
    except Exception as e:
        logger.error(f"Error generating warehouse sales PDF: {str(e)}")
        # Fallback simple PDF or error
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")

@router.get("/sales-view/excel")
async def get_warehouse_sales_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for warehouse sales.
    """
    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(WarehouseInvoice).where(
        and_(
            func.date(WarehouseInvoice.payment_date) >= from_date_obj,
            func.date(WarehouseInvoice.payment_date) <= to_date_obj
        )
    ).order_by(WarehouseInvoice.payment_date.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Sales"

    headers = ['Order ID', 'Product', 'Customer', 'Cost', 'Price', 'Qty', 'Total', 'Discount', 'Amount Paid', 'Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for inv in invoices:
        try:
            items = json.loads(inv.items) if inv.items else []
            total_invoice_paid = float(inv.amount_paid)
            
            first_item = True
            for item in items:
                product_name = item.get('product_name', 'N/A')
                quantity = item.get('quantity', 0)
                unit_price = item.get('unit_price', 0)
                item_discount = float(item.get('discount', 0))
                unit_cost = float(item.get('cost_price', 0))
                row_total = (quantity * unit_price) - item_discount
                amount_paid = total_invoice_paid if first_item else 0.0

                ws.cell(row=row_num, column=1, value=inv.invoice_no)
                ws.cell(row=row_num, column=2, value=product_name)
                ws.cell(row=row_num, column=3, value=inv.customer_name or 'N/A')
                ws.cell(row=row_num, column=4, value=unit_cost)
                ws.cell(row=row_num, column=5, value=unit_price)
                ws.cell(row=row_num, column=6, value=quantity)
                ws.cell(row=row_num, column=7, value=row_total)
                ws.cell(row=row_num, column=8, value=item_discount if first_item else 0)
                ws.cell(row=row_num, column=9, value=amount_paid)
                ws.cell(row=row_num, column=10, value=inv.payment_date.strftime('%H:%M:%S') if inv.payment_date else inv.created_at.strftime('%H:%M:%S'))
                
                row_num += 1
                first_item = False
        except Exception as e:
            logger.error(f"Error processing invoice {inv.invoice_no} for Excel: {str(e)}")
            continue

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    encoded_excel = base64.b64encode(output.read()).decode()
    return {"excel": encoded_excel, "filename": f"warehouse_sales_{from_date}_to_{to_date}.xlsx"}

@router.get("/all-invoices")
async def get_all_warehouse_invoices(
    payment_status: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Get all warehouse invoices with optional payment status filter
    """
    statement = select(WarehouseInvoice)
    
    if payment_status:
        statuses = payment_status.split(',')
        statement = statement.where(WarehouseInvoice.payment_status.in_(statuses))
    
    statement = statement.order_by(WarehouseInvoice.created_at.asc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    result_list = []
    for inv in invoices:
        try:
            totals = json.loads(inv.totals)
            balance_due = totals.get('balance_due', 0.0)
        except:
            balance_due = float(inv.total_amount - inv.amount_paid)

        result_list.append({
            "orderid": str(inv.id),
            "invoice_no": inv.invoice_no,
            "status": "completed",
            "customer": inv.customer_name,
            "teamname": "",
            "quantity": 0,
            "total_amount": float(inv.total_amount),
            "amount_paid": float(inv.amount_paid),
            "balance_due": float(balance_due),
            "payment_status": inv.payment_status,
            "date": inv.payment_date.strftime('%Y-%m-%d') if inv.payment_date else inv.created_at.strftime('%Y-%m-%d'),
            "created_at": inv.created_at.isoformat()
        })
    
    return result_list

@router.get("/customer-invoices/{customer_id}")
async def get_customer_invoices(
    customer_id: str,
    db: AsyncSession = Depends(get_db)
):
    """
    Get all invoices for a specific warehouse customer that have balance due
    """
    try:
        customer_uuid = uuid.UUID(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid customer ID format")

    statement = select(WarehouseInvoice).where(
        WarehouseInvoice.customer_id == customer_uuid
    ).order_by(WarehouseInvoice.created_at.desc())
    
    result = await db.execute(statement)
    invoices = result.scalars().all()
    
    result_list = []
    for inv in invoices:
        # Extract balance from totals JSON or use amount_paid/total_amount
        try:
            totals = json.loads(inv.totals)
            balance_due = totals.get('balance_due', 0.0)
        except:
            balance_due = float(inv.total_amount - inv.amount_paid)

        result_list.append({
            "orderid": str(inv.id),
            "invoice_no": inv.invoice_no,
            "status": "completed", # Warehouse invoices are usually completed on creation
            "customer": inv.customer_name,
            "teamname": "",
            "quantity": 0, # Could be calculated from items if needed
            "total_amount": float(inv.total_amount),
            "amount_paid": float(inv.amount_paid),
            "balance_due": float(balance_due),
            "payment_status": inv.payment_status,
            "date": inv.payment_date.strftime('%Y-%m-%d') if inv.payment_date else inv.created_at.strftime('%Y-%m-%d'),
            "created_at": inv.created_at.isoformat()
        })
    
    return result_list

@router.put("/process-payment/{order_id}")
async def process_warehouse_payment(
    order_id: str,
    payment_data: dict,
    current_user: User = Depends(admin_cashier_employee_required_from_session()),
    db: AsyncSession = Depends(get_db)
):
    """
    Process a payment for a warehouse invoice
    """
    try:
        order_uuid = uuid.UUID(order_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid order ID format")

    # Get the invoice
    statement = select(WarehouseInvoice).where(WarehouseInvoice.id == order_uuid)
    result = await db.execute(statement)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Warehouse invoice not found")

    amount = Decimal(str(payment_data.get("amount", 0)))
    payment_method = payment_data.get("payment_method", "cash")
    description = payment_data.get("description", "")
    payment_date_str = payment_data.get("payment_date")
    
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be greater than zero")

    # Update invoice fields
    try:
        totals = json.loads(invoice.totals)
    except:
        totals = {
            "subtotal": float(invoice.total_amount),
            "tax": 0.0,
            "discount": float(invoice.discounts),
            "total": float(invoice.total_amount),
            "amount_paid": float(invoice.amount_paid),
            "balance_due": float(invoice.total_amount - invoice.amount_paid),
            "payment_status": invoice.payment_status
        }

    current_balance = Decimal(str(totals.get('balance_due', 0.0)))
    
    if amount > current_balance:
        raise HTTPException(status_code=400, detail=f"Payment exceeds balance due ({current_balance})")

    new_amount_paid = invoice.amount_paid + amount
    new_balance_due = current_balance - amount
    
    # Update payment status
    if new_balance_due <= 0:
        new_status = "paid"
    else:
        new_status = "partial"

    # Add to payment history
    try:
        history = json.loads(invoice.payments_history)
    except:
        history = []
    
    payment_date = datetime.fromisoformat(payment_date_str) if payment_date_str else datetime.now()
    
    new_payment = {
        "amount": float(amount),
        "payment_method": payment_method,
        "date": payment_date.isoformat(),
        "description": description
    }
    history.append(new_payment)

    # Update totals JSON
    totals['amount_paid'] = float(new_amount_paid)
    totals['balance_due'] = float(new_balance_due)
    totals['payment_status'] = new_status

    invoice.amount_paid = new_amount_paid
    invoice.payment_status = new_status
    invoice.payments_history = json.dumps(history)
    invoice.totals = json.dumps(totals)
    invoice.updated_at = datetime.now()

    # Update WarehouseCustomer balance
    if invoice.customer_id:
        customer_stmt = select(WarehouseCustomer).where(WarehouseCustomer.id == invoice.customer_id)
        customer_result = await db.execute(customer_stmt)
        customer = customer_result.scalar_one_or_none()
        if customer:
            customer.cus_balance = (customer.cus_balance or Decimal('0')) - amount
            db.add(customer)

    db.add(invoice)
    await db.commit()
    await db.refresh(invoice)

    return {"message": "Payment processed successfully", "new_balance": float(new_balance_due)}

@router.get("/payment-history/{order_id}")
async def get_warehouse_payment_history(
    order_id: str,
    db: AsyncSession = Depends(get_db)
):
    """
    Get payment history for a warehouse invoice
    """
    try:
        order_uuid = uuid.UUID(order_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid order ID format")

    statement = select(WarehouseInvoice).where(WarehouseInvoice.id == order_uuid)
    result = await db.execute(statement)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Warehouse invoice not found")

    try:
        history = json.loads(invoice.payments_history)
    except:
        history = []
    
    try:
        totals = json.loads(invoice.totals)
        balance_due = totals.get('balance_due', 0.0)
    except:
        balance_due = float(invoice.total_amount - invoice.amount_paid)

    return {
        "invoice_id": str(invoice.id),
        "invoice_no": invoice.invoice_no,
        "total_amount": float(invoice.total_amount),
        "amount_paid": float(invoice.amount_paid),
        "balance_due": float(balance_due),
        "payment_status": invoice.payment_status,
        "payments_history": history
    }

