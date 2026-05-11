from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List
from uuid import UUID
from pydantic import BaseModel
from datetime import datetime
import logging
import base64

from ..database.database import get_db
from ..models.product import Product
from ..models.stock_entry import StockEntry, StockEntryType
from ..models.user import User
from ..models.warehouse_vendor import WarehouseVendor
from ..services.product_service import ProductService
from ..auth.session_auth import get_current_user_from_session
from sqlmodel import select
from decimal import Decimal

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/warehouse-stock", tags=["warehouse-stock"])


def warehouse_required():
    """Require warehouse, admin, or employee role from session"""
    async def role_checker(current_user: User = Depends(get_current_user_from_session)):
        if current_user.role.name not in ["warehouse", "admin", "employee"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Warehouse, admin or employee access required"
            )
        return current_user
    return role_checker


class StockInRequest(BaseModel):
    product_id: str
    qty: int
    vendor_id: Optional[str] = None
    cost_price: Optional[float] = 0.0
    ref: Optional[str] = None


class StockAdjustRequest(BaseModel):
    product_id: str
    qty: int  # positive for increase, negative for decrease
    ref: Optional[str] = None


@router.post("/in")
async def stock_in(
    request: StockInRequest,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Add stock to warehouse_stock column (products where is_warehouse_product = true)
    with optional warehouse vendor selection and balance update.
    """
    try:
        product_uuid = UUID(request.product_id)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid product ID format"
        )

    product = await ProductService.get_product(db, product_uuid)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    if not product.is_warehouse_product:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This product is not marked as warehouse product"
        )

    # Handle vendor and balance update
    vendor_id_uuid = None
    vendor_name = None
    total_cost = 0.0

    if request.vendor_id:
        try:
            vendor_id_uuid = UUID(request.vendor_id)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid vendor ID format"
            )
        
        vendor = await db.get(WarehouseVendor, vendor_id_uuid)
        if not vendor:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Warehouse vendor not found"
            )
        
        vendor_name = vendor.name
        # Update vendor balance
        total_cost = request.qty * (request.cost_price or 0.0)
        vendor.balance = (vendor.balance or Decimal('0.00')) + Decimal(str(total_cost))
        db.add(vendor)

    # Update warehouse_stock and warehouse_cost columns
    product.warehouse_stock = (product.warehouse_stock or 0) + request.qty
    product.warehouse_cost = Decimal(str(request.cost_price or 0.0))
    db.add(product)

    # Create stock entry
    stock_entry = StockEntry(
        product_id=product_uuid,
        warehouse_vendor_id=vendor_id_uuid,
        qty=request.qty,
        cost_price=Decimal(str(request.cost_price or 0.0)),
        type=StockEntryType.IN,
        location="warehouse",
        ref=request.ref or f"W_IN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    db.add(stock_entry)

    await db.commit()
    await db.refresh(product)

    return {
        "success": True,
        "message": "Stock added successfully",
        "new_stock": product.warehouse_stock,
        "vendor_name": vendor_name,
        "amount_added_to_balance": total_cost if vendor_id_uuid else None
    }


@router.post("/adjust")
async def adjust_stock(
    request: StockAdjustRequest,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Adjust warehouse_stock column (increase/decrease)
    """
    try:
        product_uuid = UUID(request.product_id)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid product ID format"
        )

    product = await ProductService.get_product(db, product_uuid)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    if not product.is_warehouse_product:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This product is not marked as warehouse product"
        )

    # Adjust warehouse_stock column
    product.warehouse_stock = (product.warehouse_stock or 0) + request.qty
    db.add(product)

    stock_entry = StockEntry(
        product_id=product_uuid,
        qty=request.qty,
        type=StockEntryType.ADJUST,
        location="warehouse",
        ref=request.ref
    )
    db.add(stock_entry)

    await db.commit()
    await db.refresh(product)

    return {
        "success": True,
        "message": "Stock adjusted successfully",
        "new_stock": product.warehouse_stock
    }


@router.get("/view")
async def view_stock(
    page: int = 1,
    limit: int = 10,
    search: str = None,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    View warehouse stock (products where is_warehouse_product = true)
    Shows warehouse_stock column values
    """
    from sqlalchemy import or_

    skip = (page - 1) * limit

    # Build query - only warehouse products with stock > 0
    base_statement = select(
        Product.id,
        Product.name,
        Product.warehouse_stock,
        Product.category,
        Product.branch,
        Product.article_no,
        Product.warehouse_limited_qty,
        Product.barcode,
        Product.sku,
        Product.unit_price,
        Product.cost_price,
        Product.is_warehouse_product,
        Product.warehouse_cost
    ).where(
        Product.is_warehouse_product == True,
        Product.warehouse_stock > 0
    )

    if search and search.strip():
        search_pattern = f"%{search.strip()}%"
        base_statement = base_statement.where(
            or_(
                Product.name.ilike(search_pattern),
                Product.barcode.ilike(search_pattern),
                Product.article_no.ilike(search_pattern),
                Product.sku.ilike(search_pattern)
            )
        )

    # Get total count - only warehouse products with stock > 0
    count_statement = select(Product.id).where(
        Product.is_warehouse_product == True,
        Product.warehouse_stock > 0
    )
    if search and search.strip():
        search_pattern = f"%{search.strip()}%"
        count_statement = count_statement.where(
            or_(
                Product.name.ilike(search_pattern),
                Product.barcode.ilike(search_pattern),
                Product.article_no.ilike(search_pattern),
                Product.sku.ilike(search_pattern)
            )
        )

    count_result = await db.execute(count_statement)
    total_count = len(count_result.scalars().all())

    # Apply pagination
    statement = base_statement.order_by(Product.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(statement)
    products = result.fetchall()

    # Targeted Vendor Fetch: Get latest warehouse vendor for ONLY these products
    product_ids = [p[0] for p in products]
    
    vendor_map = {}
    if product_ids:
        # Efficiently get the latest warehouse vendor for each product in the current page
        # Using DISTINCT ON (product_id) which is PostgreSQL specific (matching stock.py pattern)
        # Note: If not using PostgreSQL, this might need a different approach
        vendor_query = select(
            StockEntry.product_id,
            WarehouseVendor.name
        ).join(WarehouseVendor, StockEntry.warehouse_vendor_id == WarehouseVendor.id).where(
            StockEntry.product_id.in_(product_ids),
            StockEntry.location == "warehouse",
            StockEntry.type == StockEntryType.IN
        ).order_by(StockEntry.product_id, StockEntry.created_at.desc()).distinct(StockEntry.product_id)
        
        vendor_result = await db.execute(vendor_query)
        vendor_map = {row[0]: row[1] for row in vendor_result.fetchall()}

    # Format response
    result_list = [
        {
            "id": str(p[0]),
            "vendor_name": vendor_map.get(p[0], ""),
            "name": p[1],
            "warehouse_stock": p[2] or 0,
            "category": p[3] or "",
            "branch": p[4] or "",
            "article_no": p[5] or "",
            "warehouse_limited_qty": p[6] or 0,
            "barcode": p[7] or "",
            "sku": p[8] or "",
            "unit_price": float(p[9]) if p[9] else 0,
            "cost_price": float(p[10]) if p[10] else 0,
            "is_warehouse_product": p[11] or False,
            "warehouse_cost": float(p[12]) if p[12] else 0
        }
        for p in products
    ]

    total_pages = (total_count + limit - 1) // limit

    return {
        'data': result_list,
        'page': page,
        'limit': limit,
        'total': total_count,
        'total_pages': total_pages,
        'has_more': page < total_pages
    }


@router.post("/out")
async def stock_out(
    request: StockAdjustRequest,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Reduce warehouse_stock column (stock out)
    """
    try:
        product_uuid = UUID(request.product_id)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid product ID format"
        )

    product = await ProductService.get_product(db, product_uuid)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    if not product.is_warehouse_product:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This product is not marked as warehouse product"
        )

    # Check if enough stock is available
    current_stock = product.warehouse_stock or 0
    if current_stock < request.qty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Insufficient stock. Available: {current_stock}, Requested: {request.qty}"
        )

    # Reduce warehouse_stock column
    product.warehouse_stock = current_stock - request.qty
    db.add(product)

    stock_entry = StockEntry(
        product_id=product_uuid,
        qty=-request.qty,  # Negative quantity for stock out
        type=StockEntryType.OUT,
        location="warehouse",
        ref=request.ref
    )
    db.add(stock_entry)

    await db.commit()
    await db.refresh(product)

    return {
        "success": True,
        "message": "Stock removed successfully",
        "new_stock": product.warehouse_stock
    }


@router.get("/entries")
async def get_warehouse_entries(
    page: int = 1,
    limit: int = 10,
    product_id: str = None,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Get warehouse stock entries with pagination
    """
    skip = (page - 1) * limit

    # Build query
    statement = select(
        StockEntry.id,
        StockEntry.product_id,
        StockEntry.qty,
        StockEntry.type,
        StockEntry.ref,
        StockEntry.created_at,
        Product.name,
        Product.barcode
    ).join(
        Product, StockEntry.product_id == Product.id
    ).where(
        StockEntry.location == "warehouse"
    )

    if product_id:
        try:
            product_uuid = UUID(product_id)
            statement = statement.where(StockEntry.product_id == product_uuid)
        except ValueError:
            pass

    # Get total count
    count_statement = select(StockEntry.id).where(StockEntry.location == "warehouse")
    if product_id:
        try:
            product_uuid = UUID(product_id)
            count_statement = count_statement.where(StockEntry.product_id == product_uuid)
        except ValueError:
            pass

    count_result = await db.execute(count_statement)
    total_count = len(count_result.scalars().all())

    # Apply pagination
    statement = statement.order_by(StockEntry.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(statement)
    entries = result.fetchall()

    # Format response
    result_list = [
        {
            "id": str(e[0]),
            "product_id": str(e[1]),
            "qty": e[2],
            "type": e[3].value if hasattr(e[3], 'value') else str(e[3]),
            "ref": e[4] or "",
            "created_at": e[5].isoformat() if e[5] else None,
            "product_name": e[6],
            "barcode": e[7] or ""
        }
        for e in entries
    ]

    total_pages = (total_count + limit - 1) // limit

    return {
        'data': result_list,
        'page': page,
        'limit': limit,
        'total': total_count,
        'total_pages': total_pages,
        'has_more': page < total_pages
    }


@router.put("/update/{entry_id}")
async def update_stock_entry(
    entry_id: UUID,
    request: StockAdjustRequest,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Update an existing stock entry and adjust warehouse_stock accordingly
    """
    try:
        product_uuid = UUID(request.product_id)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid product ID format"
        )

    # Find the stock entry
    statement = select(StockEntry).where(StockEntry.id == entry_id)
    result = await db.execute(statement)
    stock_entry = result.scalar_one_or_none()

    if not stock_entry:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Stock entry not found"
        )

    if stock_entry.location != "warehouse":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This entry is not a warehouse stock entry"
        )

    product = await ProductService.get_product(db, product_uuid)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    if not product.is_warehouse_product:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This product is not marked as warehouse product"
        )

    # Reverse the old entry effect
    old_qty = stock_entry.qty
    product.warehouse_stock = (product.warehouse_stock or 0) - old_qty

    # Apply the new quantity
    product.warehouse_stock = (product.warehouse_stock or 0) + request.qty

    # Update the stock entry
    stock_entry.qty = request.qty
    stock_entry.ref = request.ref
    stock_entry.updated_at = datetime.now()

    db.add(product)
    db.add(stock_entry)

    await db.commit()
    await db.refresh(product)

    return {
        "success": True,
        "message": "Stock entry updated successfully",
        "new_stock": product.warehouse_stock
    }


@router.delete("/delete/{entry_id}")
async def delete_stock_entry(
    entry_id: UUID,
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Delete a stock entry and reverse its effect on warehouse_stock
    """
    # Find the stock entry
    statement = select(StockEntry).where(StockEntry.id == entry_id)
    result = await db.execute(statement)
    stock_entry = result.scalar_one_or_none()

    if not stock_entry:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Stock entry not found"
        )

    if stock_entry.location != "warehouse":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This entry is not a warehouse stock entry"
        )

    product = await ProductService.get_product(db, stock_entry.product_id)
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found"
        )

    # Reverse the stock entry effect
    product.warehouse_stock = (product.warehouse_stock or 0) - stock_entry.qty

    db.add(product)
    await db.delete(stock_entry)

    await db.commit()
    await db.refresh(product)

    return {
        "success": True,
        "message": "Stock entry deleted successfully",
        "new_stock": product.warehouse_stock
    }


@router.post("/requirement-report")
async def warehouse_requirement_report(
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate warehouse requirement report in PDF format (base64 encoded)
    Shows products that need restocking in warehouse
    Access: Warehouse, Admin
    """
    # Build query - only warehouse products
    statement = select(Product).where(Product.is_warehouse_product == True)

    result = await db.execute(statement)
    products = result.scalars().all()

    # Generate HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y')

    # Build product rows for PDF table
    product_rows = ""
    row_count = 0
    for product in products:
        # Calculate Required Stock (Limited qty - Warehouse stock)
        limited_qty = product.warehouse_limited_qty or 0
        current_stock = product.warehouse_stock or 0
        required_stock = max(0, limited_qty - current_stock)

        # Skip products that don't need restocking
        if required_stock <= 0:
            continue

        row_count += 1
        product_rows += f"""
        <tr>
            <td class="border" style="text-align: center;">{row_count}</td>
            <td class="border">{product.name}</td>
            <td class="border">{product.category or '-'}</td>
            <td class="border">{product.article_no or '-'}</td>
            <td class="border" style="text-align: right;">{float(product.cost_price) if product.cost_price else 0.0:.2f}</td>
            <td class="border" style="text-align: right;">{float(product.unit_price) if product.unit_price else 0.0:.2f}</td>
            <td class="border">{product.barcode or '-'}</td>
            <td class="border">{product.brand_action or '-'}</td>
            <td class="border" style="text-align: right;">{limited_qty}</td>
            <td class="border" style="text-align: right;">{current_stock}</td>
            <td class="border" style="text-align: right; font-weight: bold; color: #e11d48;">{required_stock}</td>
        </tr>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 10mm;
                margin-bottom: 20mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 11px;
                margin: 0;
                padding: 0;
            }}
            .signature-section {{
                margin-top: 50px;
                width: 100%;
                page-break-inside: avoid;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 22px;
                font-weight: bold;
            }}
            .print-date {{
                text-align: right;
                margin-bottom: 10px;
                color: #666;
                font-size: 10px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 5px;
            }}
            tr {{
                page-break-inside: avoid;
            }}
            th {{
                background-color: #444;
                color: white;
                border: 1px solid #000;
                padding: 8px 5px;
                text-align: left;
                font-weight: bold;
                font-size: 11px;
            }}
            td {{
                border: 1px solid #000;
                padding: 3px 5px;
                font-size: 10px;
            }}
            .border {{
                border: 1px solid #000;
            }}
            .text-right {{
                text-align: right;
            }}
            tr:nth-child(even) {{
                background-color: #f5f5f5;
            }}
            tr:nth-child(odd) {{
                background-color: #fff;
            }}
            .footer {{
                margin-top: 15px;
                text-align: center;
                font-size: 10px;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <h1>Warehouse Requirement Report</h1>
        <div class="print-date"><strong>Print Date:</strong> {current_date}</div>
        <table>
            <thead>
                <tr>
                    <th style="width: 30px;">#</th>
                    <th>Product Name</th>
                    <th>Category</th>
                    <th>Article No</th>
                    <th style="text-align: right;">Cost</th>
                    <th style="text-align: right;">Price</th>
                    <th>Barcode</th>
                    <th>Brand</th>
                    <th style="text-align: right;">Ltd Qty</th>
                    <th style="text-align: right;">Stock</th>
                    <th style="text-align: right;">Required</th>
                </tr>
            </thead>
            <tbody>
                {product_rows}
            </tbody>
        </table>
        <div class="footer">
            <p>Total Warehouse Products: {len(products)}</p>
        </div>

        <div class="signature-section">
            <table style="width: 100%; border: none !important;">
                <tr>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Prepared By
                        </div>
                    </td>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; margin-left: auto; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Approved By
                        </div>
                    </td>
                </tr>
            </table>
        </div>
    </body>
    </html>
    """

    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except ImportError:
        # Fallback to simple PDF base64 if weasyprint not available
        fallback_msg = "PDF Generation Error: weasyprint not installed"
        encoded_pdf = base64.b64encode(fallback_msg.encode()).decode()

    return encoded_pdf


@router.post("/urgent-buy-report")
async def urgent_buy_report(
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate urgent buy report (PDF base64)
    Shows products where Shop Requirement > Warehouse Stock
    """
    # Build query - only warehouse products
    statement = select(Product).where(Product.is_warehouse_product == True)
    result = await db.execute(statement)
    products = result.scalars().all()

    current_date = datetime.now().strftime('%d-%m-%Y')

    # Build product rows for PDF table
    product_rows = ""
    row_count = 0
    for product in products:
        # Shop Deficit
        shop_limit = product.limited_qty or 0
        shop_stock = product.stock_level or 0
        shop_required = max(0, shop_limit - shop_stock)

        # Warehouse Status
        warehouse_stock = product.warehouse_stock or 0
        warehouse_limit = product.warehouse_limited_qty or 0

        # Urgent Condition: Shop needs more than what warehouse has
        if shop_required > warehouse_stock:
            row_count += 1
            # How much more to buy? To cover shop requirement + maybe maintain warehouse stock
            to_buy = shop_required - warehouse_stock
            
            product_rows += f"""
            <tr>
                <td class="border" style="text-align: center;">{row_count}</td>
                <td class="border">{product.name}</td>
                <td class="border" style="text-align: center;">{shop_limit}</td>
                <td class="border" style="text-align: center;">{shop_stock}</td>
                <td class="border" style="text-align: center;">{shop_required}</td>
                <td class="border" style="text-align: center;">{warehouse_limit}</td>
                <td class="border" style="text-align: center;">{warehouse_stock}</td>
                <td class="border" style="text-align: right; font-weight: bold; color: #e11d48;">{to_buy}</td>
            </tr>
            """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 10mm; margin-bottom: 20mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 0; padding: 0; }}
            .signature-section {{ margin-top: 50px; width: 100%; page-break-inside: avoid; }}
            h1 {{ text-align: center; color: #333; margin: 0 0 10px 0; font-size: 22px; font-weight: bold; }}
            .print-date {{ text-align: right; margin-bottom: 10px; color: #666; font-size: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            tr {{ page-break-inside: avoid; }}
            th {{ background-color: #444; color: white; border: 1px solid #000; padding: 8px 5px; text-align: center; font-weight: bold; font-size: 11px; }}
            td {{ border: 1px solid #000; padding: 5px; font-size: 10px; }}
            .border {{ border: 1px solid #000; }}
            .text-right {{ text-align: right; }}
            tr:nth-child(even) {{ background-color: #f5f5f5; }}
            tr:nth-child(odd) {{ background-color: #fff; }}
        </style>
    </head>
    <body>
        <h1>Warehouse Urgent Buy Report</h1>
        <div class="print-date"><strong>Print Date:</strong> {current_date}</div>
        <table>
            <thead>
                <tr>
                    <th style="width: 30px;">#</th>
                    <th>Product Name</th>
                    <th>Shop Limit</th>
                    <th>Shop Stock</th>
                    <th>Shop Required</th>
                    <th>Warehouse Limit</th>
                    <th>Warehouse Stock</th>
                    <th style="text-align: right;">Urgent Buy Qty</th>
                </tr>
            </thead>
            <tbody>
                {product_rows if product_rows else '<tr><td colspan="8" style="text-align: center; padding: 20px;">No urgent buys required.</td></tr>'}
            </tbody>
        </table>

        <div class="signature-section">
            <table style="width: 100%; border: none !important;">
                <tr>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Verified By
                        </div>
                    </td>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; margin-left: auto; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Approved By
                        </div>
                    </td>
                </tr>
            </table>
        </div>
    </body>
    </html>
    """

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except Exception:
        fallback_msg = "PDF Generation Error"
        encoded_pdf = base64.b64encode(fallback_msg.encode()).decode()

    return encoded_pdf


@router.post("/shop-requirement-report")
async def shop_requirement_report(
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate shop requirement report for warehouse products (PDF base64)
    Shows what the shop needs based on its limited_qty and stock_level
    Access: Warehouse, Admin
    """
    # Build query - only warehouse products
    statement = select(Product).where(Product.is_warehouse_product == True)

    result = await db.execute(statement)
    products = result.scalars().all()

    # Generate HTML content for PDF
    current_date = datetime.now().strftime('%d-%m-%Y')

    # Build product rows for PDF table
    product_rows = ""
    row_count = 0
    for product in products:
        # Calculate Required Stock (Shop Limited qty - Shop Stock Level)
        limited_qty = product.limited_qty or 0
        stock_level = product.stock_level or 0
        required_stock = max(0, limited_qty - stock_level)

        # Skip products that don't need restocking
        if required_stock <= 0:
            continue

        row_count += 1
        product_rows += f"""
        <tr>
            <td class="border" style="text-align: center;">{row_count}</td>
            <td class="border">{product.name}</td>
            <td class="border">{product.category or '-'}</td>
            <td class="border">{product.barcode or '-'}</td>
            <td class="border" style="text-align: right;">{limited_qty}</td>
            <td class="border" style="text-align: right;">{stock_level}</td>
            <td class="border" style="text-align: right; font-weight: bold; color: #e11d48;">{required_stock}</td>
        </tr>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 portrait;
                margin: 15mm;
                margin-bottom: 20mm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 12px;
                margin: 0;
                padding: 0;
            }}
            .signature-section {{
                margin-top: 50px;
                width: 100%;
                page-break-inside: avoid;
            }}
            h1 {{
                text-align: center;
                color: #333;
                margin: 0 0 10px 0;
                font-size: 24px;
                font-weight: bold;
            }}
            .print-date {{
                text-align: right;
                margin-bottom: 15px;
                color: #666;
                font-size: 11px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            tr {{
                page-break-inside: avoid;
            }}
            th {{
                background-color: #444;
                color: white;
                border: 1px solid #000;
                padding: 10px 8px;
                text-align: left;
                font-weight: bold;
                font-size: 12px;
            }}
            td {{
                border: 1px solid #000;
                padding: 4px 6px;
                font-size: 11px;
            }}
            .border {{
                border: 1px solid #000;
            }}
            .text-right {{
                text-align: right;
            }}
            tr:nth-child(even) {{
                background-color: #f5f5f5;
            }}
            tr:nth-child(odd) {{
                background-color: #fff;
            }}
            .footer {{
                margin-top: 20px;
                text-align: center;
                font-size: 11px;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <h1>Shop Requirement Report</h1>
        <div class="print-date"><strong>Print Date:</strong> {current_date}</div>
        <table>
            <thead>
                <tr>
                    <th style="width: 40px;">#</th>
                    <th>Product Name</th>
                    <th>Category</th>
                    <th>Barcode</th>
                    <th style="text-align: right;">Limited Qty</th>
                    <th style="text-align: right;">Stock Level</th>
                    <th style="text-align: right;">Required</th>
                </tr>
            </thead>
            <tbody>
                {product_rows}
            </tbody>
        </table>
        <div class="footer">
            <p>Total Products: {len(products)}</p>
        </div>

        <div class="signature-section">
            <table style="width: 100%; border: none !important;">
                <tr>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Issued By
                        </div>
                    </td>
                    <td style="border: none !important; width: 50%; padding-top: 50px;">
                        <div style="width: 200px; margin-left: auto; text-align: center; font-weight: bold; font-size: 12px;">
                            <div style="border-top: 1px solid black; margin-bottom: 5px;"></div>
                            Received By
                        </div>
                    </td>
                </tr>
            </table>
        </div>
    </body>
    </html>
    """

    # Generate PDF using weasyprint
    try:
        from weasyprint import HTML
        pdf_doc = HTML(string=html_content)
        pdf_bytes = pdf_doc.write_pdf()
        encoded_pdf = base64.b64encode(pdf_bytes).decode()
    except ImportError:
        # Fallback to simple PDF base64 if weasyprint not available
        fallback_msg = "PDF Generation Error: weasyprint not installed"
        encoded_pdf = base64.b64encode(fallback_msg.encode()).decode()

    return encoded_pdf

@router.get("/adjustment-report/pdf")
async def get_warehouse_stock_adjustments_pdf(
    from_date: str = Query(...),
    to_date: str = Query(...),
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate PDF report for warehouse stock adjustments (date-wise)
    """
    import base64
    from datetime import datetime
    from sqlalchemy import and_, func
    from weasyprint import HTML

    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(StockEntry).where(
        and_(
            StockEntry.type == StockEntryType.ADJUST,
            func.date(StockEntry.created_at) >= from_date_obj,
            func.date(StockEntry.created_at) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    adjustments = result.scalars().all()
    
    product_ids = {adj.product_id for adj in adjustments}
    product_query = await db.execute(select(Product).where(Product.id.in_(product_ids)))
    products = {p.id: p for p in product_query.scalars().all()}
    
    filtered_adjustments = [
        adj for adj in adjustments 
        if products.get(adj.product_id) and products[adj.product_id].is_warehouse_product
    ]
    
    adjustment_rows = ""
    total_positive = 0
    total_negative = 0
    
    for adj in filtered_adjustments:
        product = products.get(adj.product_id)
        product_name = product.name if product else 'Unknown'
        
        adj_type = "Increase" if adj.qty > 0 else "Decrease"
        abs_qty = abs(adj.qty)
        
        if adj.qty > 0:
            total_positive += adj.qty
        else:
            total_negative += abs(adj.qty)
        
        adjustment_rows += f"""
        <tr>
            <td class="border">{adj.created_at.strftime('%Y-%m-%d')}</td>
            <td class="border">{product_name}</td>
            <td class="border">{adj_type}</td>
            <td class="border text-right">{abs_qty}</td>
            <td class="border">{adj.location or 'N/A'}</td>
            <td class="border">{adj.batch or '-'}</td>
            <td class="border">{adj.created_at.strftime('%I:%M %p')}</td>
        </tr>
        """
    
    current_date = datetime.now().strftime('%d-%m-%Y %I:%M %p')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{ size: A4 landscape; margin: 15mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 10px; }}
            h1 {{ text-align: center; color: #333; margin-bottom: 10px; font-size: 24px; }}
            .date-range {{ text-align: center; margin-bottom: 15px; color: #666; font-size: 12px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            .border {{ border: 1px solid #000; padding: 6px; }}
            th {{ background-color: #444; color: white; padding: 8px; font-size: 10px; text-align: left; }}
            .text-right {{ text-align: right; }}
            .total-row {{ background-color: #f0f0f0; font-weight: bold; }}
            .summary {{ margin-top: 20px; border: 2px solid #333; padding: 10px; }}
            .summary-row {{ display: flex; justify-content: space-between; margin-bottom: 5px; }}
        </style>
    </head>
    <body>
        <h1>Warehouse Stock Adjustment Report</h1>
        <p class="date-range">From: {from_date} To: {to_date} | Generated: {current_date}</p>
        <table>
            <thead>
                <tr>
                    <th style="width: 12%;">Date</th>
                    <th style="width: 25%;">Product</th>
                    <th style="width: 12%;">Type</th>
                    <th style="width: 10%;">Qty</th>
                    <th style="width: 15%;">Location</th>
                    <th style="width: 13%;">Batch</th>
                    <th style="width: 13%;">Time</th>
                </tr>
            </thead>
            <tbody>
                {adjustment_rows}
                <tr class="total-row">
                    <td class="border" colspan="3" style="text-align: left;">TOTAL</td>
                    <td class="border text-right">{total_positive + total_negative}</td>
                    <td class="border" colspan="3"></td>
                </tr>
            </tbody>
        </table>
        <div class="summary">
            <div class="summary-row"><span>Stock Increases:</span><span>{total_positive} units</span></div>
            <div class="summary-row"><span>Stock Decreases:</span><span>{total_negative} units</span></div>
            <div class="summary-row"><span>Net Adjustment:</span><span>{total_positive - total_negative} units</span></div>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = HTML(string=html_content).write_pdf()
    encoded_pdf = base64.b64encode(pdf_bytes).decode()

    return {"pdf": encoded_pdf}

@router.get("/adjustment-report/excel")
async def get_warehouse_stock_adjustments_excel(
    from_date: str = Query(...),
    to_date: str = Query(...),
    current_user: User = Depends(warehouse_required()),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate Excel report for warehouse stock adjustments (date-wise)
    """
    import io
    import base64
    from datetime import datetime
    from sqlalchemy import and_, func
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    try:
        from_date_obj = datetime.fromisoformat(from_date).date()
        to_date_obj = datetime.fromisoformat(to_date).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    statement = select(StockEntry).where(
        and_(
            StockEntry.type == StockEntryType.ADJUST,
            func.date(StockEntry.created_at) >= from_date_obj,
            func.date(StockEntry.created_at) <= to_date_obj
        )
    )
    result = await db.execute(statement)
    adjustments = result.scalars().all()
    
    product_ids = {adj.product_id for adj in adjustments}
    product_query = await db.execute(select(Product).where(Product.id.in_(product_ids)))
    products = {p.id: p for p in product_query.scalars().all()}
    
    filtered_adjustments = [
        adj for adj in adjustments 
        if products.get(adj.product_id) and products[adj.product_id].is_warehouse_product
    ]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Date', 'Product', 'Type', 'Qty', 'Location', 'Batch', 'Time']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, adj in enumerate(filtered_adjustments, 2):
        product = products.get(adj.product_id)
        product_name = product.name if product else 'Unknown'
        adj_type = "Increase" if adj.qty > 0 else "Decrease"
        
        ws.cell(row=row_num, column=1, value=adj.created_at.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_num, column=2, value=product_name).border = thin_border
        ws.cell(row=row_num, column=3, value=adj_type).border = thin_border
        ws.cell(row=row_num, column=4, value=abs(adj.qty)).border = thin_border
        ws.cell(row=row_num, column=5, value=adj.location or 'N/A').border = thin_border
        ws.cell(row=row_num, column=6, value=adj.batch or '-').border = thin_border
        ws.cell(row=row_num, column=7, value=adj.created_at.strftime('%I:%M %p')).border = thin_border
        
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    encoded_excel = base64.b64encode(output.read()).decode()
    
    return {"excel": encoded_excel, "filename": f"warehouse_adjustments_{from_date}_to_{to_date}.xlsx"}
